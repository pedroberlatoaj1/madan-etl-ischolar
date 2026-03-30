"""
compilador_turma.py — Converte planilha multi-abas (por turma) no formato pipeline.

Lê uma planilha gerada por gerador_planilhas.py e produz um DataFrame/XLSX
no formato esperado por cli_envio.py (1 linha = 1 aluno × 1 disciplina × 1 trimestre).

Uso:
    python compilador_turma.py --input 1A_T1_2026.xlsx --output 1A_T1_2026_pipeline.xlsx
    python compilador_turma.py --input-dir planilhas/ --output-dir pipeline/
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from avaliacao_rules import is_blank


# ---------------------------------------------------------------------------
# Leitura de metadata
# ---------------------------------------------------------------------------

def _ler_metadata(wb: openpyxl.Workbook) -> dict[str, Any]:
    """Lê a aba _metadata e retorna dict com metadados e mapa de abas."""
    if "_metadata" not in wb.sheetnames:
        raise ValueError("Aba '_metadata' não encontrada na planilha. "
                         "Este arquivo foi gerado por gerador_planilhas.py?")

    ws = wb["_metadata"]

    # Metadados gerais (linhas 2-7)
    meta: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_row=8, max_col=2, values_only=True):
        if row[0] and row[1]:
            meta[str(row[0]).strip()] = str(row[1]).strip()

    # Mapa de abas (a partir da linha 11)
    mapa_abas: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=11, max_col=5, values_only=True):
        if row[0] is None:
            break
        mapa_abas.append({
            "nome_aba": str(row[0]).strip(),
            "disciplina": str(row[1]).strip() if row[1] else "",
            "frente": str(row[2]).strip() if row[2] else "",
            "professor": str(row[3]).strip() if row[3] else "",
            "professor_nome": str(row[4]).strip() if row[4] else "",
        })

    return {
        "trimestre": meta.get("trimestre", ""),
        "turma": meta.get("turma", ""),
        "serie": meta.get("serie", ""),
        "ano": meta.get("ano", ""),
        "gerado_em": meta.get("gerado_em", ""),
        "mapa_abas": mapa_abas,
    }


# ---------------------------------------------------------------------------
# Compilação de uma aba
# ---------------------------------------------------------------------------

def _formatar_frente_professor(frente: str, professor: str) -> str:
    """
    Formata o campo 'Frente - Professor' no formato esperado pelo pipeline.

    Exemplos:
        ("F2", "Luan") → "F2 - Luan"
        ("", "Nery")   → "Nery"
    """
    if frente:
        return f"{frente} - {professor}"
    return professor


def _compilar_aba(
    wb: openpyxl.Workbook,
    aba_info: dict[str, str],
    trimestre: str,
) -> list[dict[str, Any]]:
    """
    Lê uma aba de disciplina e retorna linhas no formato pipeline.

    Pula linhas onde TODAS as colunas de nota estão em branco.
    """
    nome_aba = aba_info["nome_aba"]
    if nome_aba not in wb.sheetnames:
        return []

    ws = wb[nome_aba]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    frente_professor = _formatar_frente_professor(
        aba_info["frente"], aba_info["professor"]
    )

    linhas: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        row_dict = dict(zip(headers, row_values))

        nome = row_dict.get("Nome", "")
        ra = row_dict.get("RA", "")
        turma = row_dict.get("Turma", "")

        if is_blank(nome) or is_blank(ra):
            continue

        # Verifica se há pelo menos uma nota preenchida
        colunas_nota = [
            "AV 1 (OBJ)", "AV 1 (DISC)",
            "AV 2 (OBJ)", "AV 2 (DISC)",
            "AV 3 (listas)", "AV 3 (avaliação)",
            "Simulado", "Ponto Extra", "Recuperação",
        ]
        tem_nota = any(not is_blank(row_dict.get(col)) for col in colunas_nota)
        if not tem_nota:
            continue

        # Emite linha no formato pipeline
        linha: dict[str, Any] = {
            "Estudante": nome,
            "RA": ra,
            "Turma": turma,
            "Trimestre": trimestre,
            "Disciplina": aba_info["disciplina"],
            "Frente - Professor": frente_professor,
        }

        # Colunas de nota (preserva valores originais)
        for col in colunas_nota + ["Obs Ponto Extra"]:
            val = row_dict.get(col)
            linha[col] = val if not is_blank(val) else None

        # Colunas de conferência
        for col in ["Nota sem a AV 3", "Nota com a AV 3", "Nota Final"]:
            val = row_dict.get(col)
            linha[col] = val if not is_blank(val) else None

        linhas.append(linha)

    return linhas


# ---------------------------------------------------------------------------
# Compilação de arquivo completo
# ---------------------------------------------------------------------------

def compilar_planilha_turma(input_path: str | Path) -> pd.DataFrame:
    """
    Lê uma planilha multi-abas e retorna DataFrame no formato pipeline.

    O DataFrame resultante pode ser salvo como XLSX/CSV e consumido por
    cli_envio.py sem nenhuma mudança no pipeline.
    """
    path = Path(input_path)
    wb = openpyxl.load_workbook(str(path), data_only=True)

    metadata = _ler_metadata(wb)
    trimestre = metadata["trimestre"]

    if not trimestre:
        raise ValueError(f"Trimestre não encontrado na metadata de {path}")

    todas_linhas: list[dict[str, Any]] = []
    for aba_info in metadata["mapa_abas"]:
        linhas = _compilar_aba(wb, aba_info, trimestre)
        todas_linhas.extend(linhas)

    wb.close()

    if not todas_linhas:
        return pd.DataFrame()

    df = pd.DataFrame(todas_linhas)

    # Ordena por aluno, depois disciplina
    df = df.sort_values(["Estudante", "Disciplina"]).reset_index(drop=True)

    return df


def compilar_planilha_para_arquivo(
    input_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Compila uma planilha multi-abas e salva como XLSX pipeline-compatível."""
    df = compilar_planilha_turma(input_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(str(out), index=False)
    return out


def compilar_diretorio(
    input_dir: str | Path,
    output_dir: str | Path,
) -> list[Path]:
    """Compila todos os XLSX de um diretório."""
    in_path = Path(input_dir)
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    arquivos_gerados: list[Path] = []
    for xlsx in sorted(in_path.glob("*.xlsx")):
        if xlsx.name.startswith("~"):  # Ignora arquivos temporários do Excel
            continue
        output_name = xlsx.stem + "_pipeline.xlsx"
        output_file = out_path / output_name
        try:
            compilar_planilha_para_arquivo(xlsx, output_file)
            arquivos_gerados.append(output_file)
        except ValueError as e:
            print(f"Pulando {xlsx.name}: {e}")

    return arquivos_gerados


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compila planilhas multi-abas para formato pipeline."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input", help="Caminho de uma planilha multi-abas")
    group.add_argument("--input-dir", help="Diretório com planilhas multi-abas")

    parser.add_argument("--output", help="Caminho do arquivo de saída (com --input)")
    parser.add_argument("--output-dir", default="./pipeline",
                        help="Diretório de saída (com --input-dir)")
    args = parser.parse_args()

    if args.input:
        output = args.output or str(
            Path(args.input).stem + "_pipeline.xlsx"
        )
        compilar_planilha_para_arquivo(args.input, output)
        print(f"Compilado: {args.input} → {output}")
    else:
        arquivos = compilar_diretorio(args.input_dir, args.output_dir)
        print(f"\n{len(arquivos)} arquivo(s) compilado(s):")
        for arq in arquivos:
            print(f"  {arq}")


if __name__ == "__main__":
    main()
