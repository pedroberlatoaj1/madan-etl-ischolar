"""
gerador_planilhas.py — Gera planilhas Excel no formato wide (1 linha por aluno).

MODO PADRÃO (por turma):
    Gera 1 arquivo por turma com uma única aba "Notas":
    - 4 colunas fixas pré-preenchidas: Estudante, RA, Turma, Trimestre
    - Colunas dinâmicas por (disciplina, frente, tipo de avaliação):
          "{Disciplina} - Frente {X} - {Tipo}"
      ex: "Matemática - Frente A - AV 1 Obj"

    Uso:
        python gerador_planilhas.py --trimestre T1 --ano 2026 --alunos roster.csv --output ./planilhas/

MODO ANUAL (workbook único — Plano B):
    Gera 1 workbook com 12 abas trimestrais:
        1A_T1, 1A_T2, 1A_T3
        1B_T1, 1B_T2, 1B_T3
        2A_T1, 2A_T2, 2A_T3
        2B_T1, 2B_T2, 2B_T3

    Cada aba tem o mesmo formato wide homologado.
    O pedagógico usa 1 arquivo por ano; o pipeline processa 1 aba por vez.

    Uso:
        python gerador_planilhas.py --anual --ano 2026 --alunos roster.csv --output ./planilhas/

O output é consumido diretamente pelo wide_format_adapter.py no pipeline ETL.
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from madan_planilha_mapper import SERIES_SUPORTADAS, extrair_serie_da_turma
from professores_madan import (
    PROFESSORES,
    SIGLA_PARA_DISCIPLINA,
    buscar_professor_para_turma,
)


# ---------------------------------------------------------------------------
# Constantes — formato antigo (mantidas para compatibilidade)
# ---------------------------------------------------------------------------

BULLET = "\u2022"  # "•" — indica "todas as turmas/frentes"

COLUNAS_IDENTIDADE = ["Nome", "RA", "Turma"]

COLUNAS_NOTA = [
    "AV 1 (OBJ)",
    "AV 1 (DISC)",
    "AV 2 (OBJ)",
    "AV 2 (DISC)",
    "AV 3 (listas)",
    "AV 3 (avaliação)",
    "Simulado",
    "Ponto Extra",
    "Obs Ponto Extra",
    "Recuperação",
]

COLUNAS_CONFERENCIA = [
    "Nota sem a AV 3",
    "Nota com a AV 3",
    "Nota Final",
]

TODAS_COLUNAS = COLUNAS_IDENTIDADE + COLUNAS_NOTA + COLUNAS_CONFERENCIA


# ---------------------------------------------------------------------------
# Constantes — formato wide novo
# ---------------------------------------------------------------------------

DISCIPLINA_DISPLAY: dict[str, str] = {
    "arte":                   "Arte",
    "biologia":               "Biologia",
    "educacao fisica":        "Educação Física",
    "filosofia":              "Filosofia",
    "fisica":                 "Física",
    "geografia":              "Geografia",
    "gramatica":              "Gramática",
    "historia":               "História",
    "ingles":                 "Inglês",
    "interpretacao de texto": "Interpretação de Texto",
    "literatura":             "Literatura",
    "matematica":             "Matemática",
    "quimica":                "Química",
    "redacao":                "Redação",
    "sociologia":             "Sociologia",
    "xadrez":                 "Xadrez",
}
"""Mapa slug canônico → nome de exibição na planilha."""

TIPOS_AVALIACAO_WIDE: list[str] = [
    "AV 1 Obj",
    "AV 1 Disc",
    "AV 2 Obj",
    "AV 2 Disc",
    "AV 3 Listas",
    "AV 3 Avaliacao",
    "Simulado",
    "Ponto Extra",
    "Recuperação",
]
"""
Sufixos de tipo de avaliação para as colunas dinâmicas.

Cada string, ao ser normalizada por wide_format_adapter._normalizar_texto(),
deve bater exatamente com uma chave de MAPA_TIPO_AVALIACAO:
    "AV 1 Obj"    → "av 1 obj"    → "AV 1 (OBJ)"
    "AV 1 Disc"   → "av 1 disc"   → "AV 1 (DISC)"
    "AV 3 Listas" → "av 3 listas" → "AV 3 (listas)"
    "AV 3 Avaliacao" → "av 3 avaliacao" → "AV 3 (avaliação)"
    "Recuperação" → "recuperacao" → "Recuperação"
"""

COLUNAS_FIXAS_WIDE = ["Estudante", "RA", "Turma", "Trimestre"]
"""Colunas fixas do formato wide — devem ser as 4 primeiras."""


# ---------------------------------------------------------------------------
# Dataclass TabConfig — mantida para retrocompatibilidade com testes existentes
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class TabConfig:
    """Representa uma combinação disciplina-frente-professor (formato legado)."""
    disciplina: str
    frente: str
    professor_display: str
    professor_nome: str

    @property
    def nome_aba(self) -> str:
        if self.frente:
            raw = f"{self.disciplina}_{self.frente}_{self.professor_display}"
        else:
            raw = f"{self.disciplina}_{self.professor_display}"
        return _sanitizar_nome_aba(raw)[:31]


def _sanitizar_nome_aba(nome: str) -> str:
    """Remove acentos e caracteres proibidos em nomes de aba do Excel."""
    nfkd = unicodedata.normalize("NFD", nome)
    sem_acento = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    return re.sub(r'[\\/*?\[\]:]', '_', sem_acento)


# ---------------------------------------------------------------------------
# Descoberta de abas (formato legado — mantida)
# ---------------------------------------------------------------------------

def descobrir_tabs_para_turma(serie: int, turma_letra: str) -> list[TabConfig]:
    """
    Retorna lista de TabConfig (formato legado) para a turma.
    Mantida para retrocompatibilidade.
    """
    disciplinas = sorted(set(SIGLA_PARA_DISCIPLINA.values()))
    tabs: set[tuple[str, str, str, str]] = set()

    for disc in disciplinas:
        profs = buscar_professor_para_turma(disc, serie, turma_letra)
        for p in profs:
            all_frentes = p.frentes_med + p.frentes_ext + p.frentes_ita
            if not all_frentes:
                continue
            has_bullet = BULLET in all_frentes
            if has_bullet:
                tabs.add((disc, "", p.nome_display, p.nome))
            else:
                expanded: set[str] = set()
                for f in all_frentes:
                    for sub in f.split("/"):
                        s = sub.strip()
                        if s and s != BULLET:
                            expanded.add(s)
                for fr in expanded:
                    tabs.add((disc, fr, p.nome_display, p.nome))

    return sorted(
        [TabConfig(d, f, pd, pn) for d, f, pd, pn in tabs],
        key=lambda t: (t.disciplina, t.frente, t.professor_display),
    )


# ---------------------------------------------------------------------------
# Descoberta e construção de colunas — formato wide novo
# ---------------------------------------------------------------------------

def descobrir_grupos_wide(serie: int, turma_letra: str) -> list[tuple[str, str]]:
    """
    Retorna lista ordenada de (disciplina_display, frente_display) para a turma.

    Regras de mapeamento de frentes:
    - Disciplina com 1 entrada (frente="" bullet ou único código) → "Frente Única"
    - Disciplina com N entradas → "Frente A", "Frente B", ... (ordem alfabética dos códigos)

    Os nomes gerados são compatíveis com parsear_coluna_dinamica() e com as
    chaves de mapa_professores.json via construir_frente_professor().
    """
    tabs = descobrir_tabs_para_turma(serie, turma_letra)

    por_disc: dict[str, list[str]] = {}
    for tab in tabs:
        por_disc.setdefault(tab.disciplina, []).append(tab.frente)

    grupos: list[tuple[str, str]] = []
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    for disc_slug in sorted(por_disc.keys()):
        frentes_raw = sorted(set(por_disc[disc_slug]))
        disc_display = DISCIPLINA_DISPLAY.get(disc_slug, disc_slug.capitalize())

        if len(frentes_raw) == 1:
            # Um único professor/frente → Frente Única
            grupos.append((disc_display, "Frente Única"))
        else:
            # Múltiplas frentes → ordenar e atribuir letras A, B, C...
            for i in range(len(frentes_raw)):
                grupos.append((disc_display, f"Frente {letras[i]}"))

    return grupos


def construir_cabecalho_wide(grupos: list[tuple[str, str]]) -> list[str]:
    """
    Constrói a lista completa de nomes de colunas para o formato wide.

    Estrutura:
        ["Estudante", "RA", "Turma", "Trimestre",
         "Disciplina X - Frente A - AV 1 Obj",
         "Disciplina X - Frente A - AV 1 Disc",
         ...]

    Os nomes dinâmicos são gerados no padrão exato que parsear_coluna_dinamica()
    em wide_format_adapter.py reconhece.
    """
    cabecalho = list(COLUNAS_FIXAS_WIDE)
    for disc_display, frente_display in grupos:
        for tipo in TIPOS_AVALIACAO_WIDE:
            cabecalho.append(f"{disc_display} - {frente_display} - {tipo}")
    return cabecalho


# ---------------------------------------------------------------------------
# Leitura de roster de alunos
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Aluno:
    """Dados base de um aluno para pré-preenchimento."""
    nome: str
    ra: str
    turma: str


def carregar_roster_csv(caminho: str | Path) -> list[Aluno]:
    """
    Carrega roster de alunos de um CSV.

    Espera colunas: Nome (ou Estudante), RA, Turma (case-insensitive).
    """
    path = Path(caminho)
    alunos: list[Aluno] = []

    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"CSV vazio ou sem cabeçalho: {path}")

        header_map: dict[str, str] = {}
        for field in reader.fieldnames:
            low = field.strip().lower()
            if low in ("nome", "estudante", "aluno", "nome_aluno"):
                header_map["nome"] = field
            elif low in ("ra", "registro_aluno", "numero_re"):
                header_map["ra"] = field
            elif low in ("turma", "sala", "classe"):
                header_map["turma"] = field

        for key in ("nome", "ra", "turma"):
            if key not in header_map:
                raise ValueError(
                    f"Coluna '{key}' não encontrada no CSV. "
                    f"Colunas disponíveis: {reader.fieldnames}"
                )

        for row in reader:
            nome = (row.get(header_map["nome"]) or "").strip()
            ra = (row.get(header_map["ra"]) or "").strip()
            turma = (row.get(header_map["turma"]) or "").strip()
            if nome and ra:
                alunos.append(Aluno(nome=nome, ra=ra, turma=turma))

    return alunos


def agrupar_alunos_por_turma(alunos: list[Aluno]) -> dict[str, list[Aluno]]:
    """Agrupa alunos por turma, retornando dict ordenado."""
    grupos: dict[str, list[Aluno]] = {}
    for a in alunos:
        grupos.setdefault(a.turma, []).append(a)
    for turma in grupos:
        grupos[turma].sort(key=lambda x: x.nome)
    return dict(sorted(grupos.items()))


# ---------------------------------------------------------------------------
# Estilos Excel
# ---------------------------------------------------------------------------

_FILL_HEADER    = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_FILL_FIXAS     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FONT_HEADER    = Font(bold=True, color="FFFFFF", size=10, name="Arial")
_FONT_DADOS     = Font(size=10, name="Arial")
_ALIGN_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center")


# ---------------------------------------------------------------------------
# Geração da aba wide
# ---------------------------------------------------------------------------

def _criar_aba_notas_wide(
    wb: openpyxl.Workbook,
    cabecalho: list[str],
    alunos: list[Aluno],
    turma: str,
    trimestre: str,
    titulo: str = "Notas",
) -> None:
    """
    Cria uma aba no formato wide:
    - Linha 1: cabeçalho completo
    - Linhas 2+: 1 aluno por linha, colunas fixas pré-preenchidas, notas em branco

    O parâmetro `titulo` define o nome da aba (default "Notas" para modo por turma;
    no modo anual recebe "1A_T1", "2B_T3", etc.).
    """
    ws = wb.create_sheet(title=titulo)
    n_fixas = len(COLUNAS_FIXAS_WIDE)

    # --- Cabeçalho ---
    for c_idx, col in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER

    # --- Dados dos alunos ---
    for r_idx, aluno in enumerate(alunos, 2):
        valores_fixos = [aluno.nome, aluno.ra, turma, trimestre]
        for c_idx, val in enumerate(valores_fixos, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = _FILL_FIXAS
            cell.font = _FONT_DADOS
            cell.alignment = _ALIGN_LEFT if c_idx == 1 else _ALIGN_CENTER
        # Colunas de nota ficam em branco (professores preenchem)

    # --- Larguras ---
    ws.column_dimensions["A"].width = 30   # Estudante
    ws.column_dimensions["B"].width = 8    # RA
    ws.column_dimensions["C"].width = 8    # Turma
    ws.column_dimensions["D"].width = 11   # Trimestre
    for c in range(n_fixas + 1, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # --- Altura do cabeçalho ---
    ws.row_dimensions[1].height = 55

    # --- Congelar colunas fixas + linha de cabeçalho ---
    ws.freeze_panes = "E2"


# ---------------------------------------------------------------------------
# Geração de planilha por turma (entrypoint principal)
# ---------------------------------------------------------------------------

def gerar_planilha_turma(
    turma: str,
    trimestre: str,
    ano: int,
    alunos: list[Aluno],
    output_dir: str | Path,
) -> Path:
    """
    Gera um arquivo Excel para uma turma no formato wide (1 aba "Notas").

    O cabeçalho é construído dinamicamente a partir do registro de professores
    da turma, garantindo que as colunas geradas sejam reconhecidas pelo
    wide_format_adapter.py no pipeline ETL.

    Retorna o caminho do arquivo gerado.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    serie = extrair_serie_da_turma(turma)
    if serie is None:
        raise ValueError(f"Não foi possível extrair série da turma: {turma!r}")
    if serie not in SERIES_SUPORTADAS:
        raise ValueError(
            f"Série {serie} (turma {turma!r}) não é suportada. "
            f"Séries suportadas: {SERIES_SUPORTADAS}"
        )

    # Extrai letra da turma
    turma_letra = ""
    for c in str(turma).strip():
        if c.isalpha():
            turma_letra = c.upper()
            break
    if not turma_letra:
        raise ValueError(f"Não foi possível extrair letra da turma: {turma!r}")

    grupos = descobrir_grupos_wide(serie, turma_letra)
    if not grupos:
        raise ValueError(
            f"Nenhuma combinação disciplina-frente encontrada para "
            f"turma {turma!r} (série {serie}, letra {turma_letra})"
        )

    cabecalho = construir_cabecalho_wide(grupos)

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _criar_aba_notas_wide(wb, cabecalho, alunos, turma, trimestre)

    filename = f"{turma}_{trimestre}_{ano}.xlsx"
    filepath = output_path / filename
    wb.save(str(filepath))

    return filepath


def gerar_todas_planilhas(
    trimestre: str,
    ano: int,
    alunos: list[Aluno],
    output_dir: str | Path,
) -> list[Path]:
    """
    Gera planilhas wide para todas as turmas presentes no roster.

    Retorna lista de caminhos dos arquivos gerados.
    """
    grupos = agrupar_alunos_por_turma(alunos)
    arquivos: list[Path] = []

    for turma, alunos_turma in grupos.items():
        serie = extrair_serie_da_turma(turma)
        if serie is None or serie not in SERIES_SUPORTADAS:
            continue

        filepath = gerar_planilha_turma(
            turma=turma,
            trimestre=trimestre,
            ano=ano,
            alunos=alunos_turma,
            output_dir=output_dir,
        )
        arquivos.append(filepath)

    return arquivos


# ---------------------------------------------------------------------------
# Modo Anual (Plano B) — workbook único com 12 abas trimestrais
# ---------------------------------------------------------------------------

#: Turmas suportadas no workbook anual, em ordem de exibição.
TURMAS_ANUAIS: list[str] = ["1A", "1B", "2A", "2B"]

#: Trimestres em ordem cronológica.
TRIMESTRES: list[str] = ["T1", "T2", "T3"]


def gerar_workbook_anual(
    ano: int,
    alunos_por_turma: dict[str, list[Aluno]],
    output_dir: str | Path,
    turmas: list[str] | None = None,
    trimestres: list[str] | None = None,
) -> Path:
    """
    Gera um único workbook anual com N abas trimestrais no formato wide.

    Estrutura de abas gerada (padrão):
        1A_T1, 1A_T2, 1A_T3,
        1B_T1, 1B_T2, 1B_T3,
        2A_T1, 2A_T2, 2A_T3,
        2B_T1, 2B_T2, 2B_T3

    Cada aba tem o mesmo formato wide homologado de `gerar_planilha_turma`:
    - Linha 1: cabeçalho completo (colunas fixas + dinâmicas por disciplina/frente/tipo)
    - Linhas 2+: 1 aluno por linha, colunas fixas pré-preenchidas, notas em branco
    - Freeze em E2 (cabeçalho + 4 colunas fixas)

    O pipeline processa 1 aba por vez — compatível com wide_format_adapter.py.

    Parâmetros:
        ano              : Ano letivo (ex: 2026)
        alunos_por_turma : dict {turma → lista de Aluno}; turmas ausentes → aba sem alunos
        output_dir       : Diretório de saída
        turmas           : Quais turmas incluir (default: TURMAS_ANUAIS = 1A,1B,2A,2B)
        trimestres       : Quais trimestres incluir (default: TRIMESTRES = T1,T2,T3)

    Retorna o caminho do arquivo gerado.
    """
    turmas_efetivas = [t.upper() for t in (turmas or TURMAS_ANUAIS)]
    trimestres_efetivos = [t.upper() for t in (trimestres or TRIMESTRES)]

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    abas_geradas: list[str] = []

    for turma in turmas_efetivas:
        # Extrai série e letra da turma (ex: "1A" → serie=1, letra="A")
        serie = extrair_serie_da_turma(turma)
        if serie is None or serie not in SERIES_SUPORTADAS:
            continue

        turma_letra = ""
        for c in turma:
            if c.isalpha():
                turma_letra = c.upper()
                break
        if not turma_letra:
            continue

        grupos = descobrir_grupos_wide(serie, turma_letra)
        if not grupos:
            continue

        cabecalho = construir_cabecalho_wide(grupos)
        alunos_turma = alunos_por_turma.get(turma, [])

        for trimestre in trimestres_efetivos:
            titulo_aba = f"{turma}_{trimestre}"          # ex: "1A_T1"
            _criar_aba_notas_wide(
                wb=wb,
                cabecalho=cabecalho,
                alunos=alunos_turma,
                turma=turma,
                trimestre=trimestre,
                titulo=titulo_aba,
            )
            abas_geradas.append(titulo_aba)

    if not abas_geradas:
        raise ValueError(
            "Nenhuma aba foi gerada. Verifique se as turmas são suportadas "
            f"({SERIES_SUPORTADAS}) e se os grupos foram descobertos."
        )

    filename = f"madan_{ano}_anual.xlsx"
    filepath = output_path / filename
    wb.save(str(filepath))

    return filepath


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Gera planilhas Excel no formato wide.\n"
            "Modo padrão: 1 arquivo por turma/trimestre.\n"
            "Modo anual (--anual): 1 workbook com 12 abas trimestrais (Plano B)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--ano", type=int, required=True, help="Ano letivo (ex: 2026)")
    parser.add_argument("--alunos", required=True, help="Caminho do CSV com roster de alunos")
    parser.add_argument("--output", default="./planilhas", help="Diretório de saída")

    # Modo padrão: trimestre obrigatório
    parser.add_argument(
        "--trimestre",
        default=None,
        help="Trimestre (T1, T2, T3) — obrigatório no modo padrão, ignorado com --anual",
    )

    # Modo anual
    parser.add_argument(
        "--anual",
        action="store_true",
        default=False,
        help=(
            "Gera workbook anual único com 12 abas (1A_T1 … 2B_T3). "
            "Não requer --trimestre."
        ),
    )

    args = parser.parse_args()

    alunos = carregar_roster_csv(args.alunos)
    print(f"Roster carregado: {len(alunos)} alunos")

    if args.anual:
        # --- Modo anual (Plano B) ---
        alunos_por_turma = agrupar_alunos_por_turma(alunos)
        filepath = gerar_workbook_anual(
            ano=args.ano,
            alunos_por_turma=alunos_por_turma,
            output_dir=args.output,
        )
        print(f"\nWorkbook anual gerado:")
        print(f"  {filepath}")
        n_abas = len(TURMAS_ANUAIS) * len(TRIMESTRES)
        print(f"  {n_abas} abas: {', '.join(f'{t}_{tr}' for t in TURMAS_ANUAIS for tr in TRIMESTRES)}")

    else:
        # --- Modo padrão: por turma/trimestre ---
        if not args.trimestre:
            parser.error("--trimestre é obrigatório no modo padrão (ou use --anual)")

        arquivos = gerar_todas_planilhas(
            trimestre=args.trimestre.upper(),
            ano=args.ano,
            alunos=alunos,
            output_dir=args.output,
        )

        print(f"\n{len(arquivos)} planilha(s) gerada(s):")
        for arq in arquivos:
            print(f"  {arq}")


if __name__ == "__main__":
    main()
