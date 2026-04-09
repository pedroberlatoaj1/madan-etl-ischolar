"""
verificar_cadastro.py - Verificacao standalone de cadastro de alunos no iScholar.

Objetivo:
- ler uma planilha multi-aba com alunos por turma;
- verificar se cada RA resolve aluno + matricula no iScholar;
- reportar falhas de cadastro sem depender do pipeline de notas.

Este utilitario nao envia notas, nao depende de trimestre/disciplinas e nao
altera o pipeline oficial.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

import openpyxl

from ischolar_client import IScholarClient
from madan_planilha_mapper import CAN_ESTUDANTE, CAN_RA, mapear_colunas_madan
from resolvedor_ids_ischolar import _extrair_id_aluno_da_resposta


_RE_ABA_PLANO_B = re.compile(
    r"^(?P<turma>[1-9][A-Za-z])_(?P<trimestre>T[123])$",
    re.IGNORECASE,
)


def parsear_nome_aba_plano_b(nome_aba: str) -> Optional[tuple[str, str]]:
    """
    Interpreta o nome de uma aba como aba trimestral Plano B.

    Retorna (turma, trimestre) se seguir o padrão, None caso contrário.

    Exemplos:
        "2A_T1" → ("2A", "T1")
        "1B_T3" → ("1B", "T3")
        "Notas"  → None
        "1A"     → None
    """
    m = _RE_ABA_PLANO_B.match(str(nome_aba).strip())
    if not m:
        return None
    return m.group("turma").upper(), m.group("trimestre").upper()


def deduplicar_turmas_plano_b(
    turmas: dict[str, list[dict[str, str]]],
) -> tuple[dict[str, list[dict[str, str]]], dict[str, list[str]], list[str]]:
    """
    Agrupa abas Plano B (ex: 2A_T1, 2A_T2, 2A_T3) pela turma base, eliminando
    RAs duplicados entre trimestres.

    Abas que não seguem o padrão Plano B passam inalteradas no resultado.

    Retorna:
        turmas_result  — dict[turma_ou_aba, list[aluno]] com dedup aplicado
        fontes_result  — dict[turma, list[nome_aba]] indicando origem de cada turma merged
        avisos         — mensagens informativas sobre a dedup

    Regras:
    - Um RA visto em T1 que reaparece em T2/T3 é ignorado nas abas seguintes.
    - A ordem entre trimestres é preservada (T1 → T2 → T3).
    - Abas sem padrão Plano B coexistem sem interferência.
    """
    plano_b: dict[str, dict] = {}
    passthrough: dict[str, list[dict[str, str]]] = {}

    for nome_aba, alunos in turmas.items():
        parsed = parsear_nome_aba_plano_b(nome_aba)
        if parsed is None:
            passthrough[nome_aba] = alunos
        else:
            turma, trimestre = parsed
            if turma not in plano_b:
                plano_b[turma] = {"alunos": [], "fontes": [], "ras_vistos": set()}
            bucket = plano_b[turma]
            bucket["fontes"].append(nome_aba)
            for aluno in alunos:
                if aluno["ra"] not in bucket["ras_vistos"]:
                    bucket["ras_vistos"].add(aluno["ra"])
                    bucket["alunos"].append(aluno)

    turmas_result: dict[str, list[dict[str, str]]] = dict(passthrough)
    fontes_result: dict[str, list[str]] = {}
    avisos: list[str] = []

    for turma, bucket in sorted(plano_b.items()):
        fontes_ordenadas = sorted(bucket["fontes"])
        turmas_result[turma] = bucket["alunos"]
        fontes_result[turma] = fontes_ordenadas
        n_abas = len(fontes_ordenadas)
        n_alunos = len(bucket["alunos"])
        if n_abas > 1:
            avisos.append(
                f"Turma '{turma}': {n_alunos} alunos unicos agregados de "
                f"{n_abas} abas ({', '.join(fontes_ordenadas)})"
            )

    return turmas_result, fontes_result, avisos


@dataclass
class ResultadoAluno:
    turma: str
    ra: str
    nome: str
    status: str
    id_aluno: Optional[int] = None
    id_matricula: Optional[int] = None
    detalhe: str = ""
    heuristica_usada: str = ""


@dataclass
class RelatorioTurma:
    turma: str
    total: int = 0
    ok: int = 0
    heuristica: int = 0
    falhas: int = 0
    alunos: list[ResultadoAluno] = field(default_factory=list)
    fontes: list[str] = field(default_factory=list)
    """Abas de origem quando deduplicated em modo Plano B (ex: ["2A_T1","2A_T2","2A_T3"])."""


@dataclass
class LeituraPlanilha:
    turmas: dict[str, list[dict[str, str]]]
    avisos: list[str] = field(default_factory=list)


class LeituraPlanilhaError(ValueError):
    """Erro de leitura defensiva da planilha de cadastro."""


def _coagir_ra(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return ""
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else str(valor).strip()

    texto = str(valor).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        return texto[:-2]
    return texto


def _parece_ra(valor: Any) -> bool:
    texto = _coagir_ra(valor)
    return bool(texto) and texto.isdigit()


def _parece_nome(valor: Any) -> bool:
    if valor is None:
        return False
    texto = str(valor).strip()
    return bool(texto) and any(ch.isalpha() for ch in texto)


def _extrair_celula(row: tuple[Any, ...], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _resolver_colunas_planilha(
    headers: list[Any],
    amostras: list[tuple[Any, ...]],
    nome_aba: str,
) -> tuple[int, int, Optional[str]]:
    headers_txt = [str(h).strip() if h is not None else "" for h in headers]
    mapping = mapear_colunas_madan(headers_txt)

    idx_estudante = [i for i, h in enumerate(headers_txt) if mapping.get(h) == CAN_ESTUDANTE]
    idx_ra = [i for i, h in enumerate(headers_txt) if mapping.get(h) == CAN_RA]

    if len(idx_estudante) != 1 or len(idx_ra) != 1:
        raise LeituraPlanilhaError(
            f"Aba '{nome_aba}' deve ter exatamente 1 coluna de nome/aluno e 1 coluna de RA. "
            f"Encontrado: estudante={len(idx_estudante)}, ra={len(idx_ra)}."
        )

    col_estudante = idx_estudante[0]
    col_ra = idx_ra[0]

    sinais_normais = 0
    sinais_invertidos = 0
    for row in amostras:
        valor_est = _extrair_celula(row, col_estudante)
        valor_ra = _extrair_celula(row, col_ra)
        if valor_est is None and valor_ra is None:
            continue

        if _parece_nome(valor_est) and _parece_ra(valor_ra):
            sinais_normais += 1
        elif _parece_ra(valor_est) and _parece_nome(valor_ra):
            sinais_invertidos += 1

    if sinais_normais and sinais_invertidos:
        raise LeituraPlanilhaError(
            f"Aba '{nome_aba}' tem dados inconsistentes: parte das linhas indica "
            "colunas corretas e parte indica colunas invertidas."
        )

    if sinais_invertidos and not sinais_normais:
        aviso = (
            f"Aba '{nome_aba}' esta com os valores de nome e RA invertidos em relacao "
            "ao cabecalho; leitura ajustada automaticamente."
        )
        return col_ra, col_estudante, aviso

    return col_estudante, col_ra, None


def ler_planilha_multi_aba(
    caminho: str,
    aba_filtro: Optional[str] = None,
) -> LeituraPlanilha:
    """
    Le por cabecalho real da linha 1 e extrai Nome/RA de cada aba.

    Regras defensivas:
    - suporta aliases de cabecalho do projeto (Estudante/Aluno/Nome e RA);
    - detecta inversao consistente entre nome e RA e corrige com aviso;
    - rejeita situacoes ambiguas ou inconsistentes.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    turmas: dict[str, list[dict[str, str]]] = {}
    avisos: list[str] = []

    for nome_aba in wb.sheetnames:
        if aba_filtro and nome_aba.strip().upper() != aba_filtro.strip().upper():
            continue

        ws = wb[nome_aba]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headers:
            continue

        amostras = list(ws.iter_rows(min_row=2, max_row=12, values_only=True))
        idx_nome, idx_ra, aviso = _resolver_colunas_planilha(list(headers), amostras, nome_aba)
        if aviso:
            avisos.append(aviso)

        alunos: list[dict[str, str]] = []
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nome_val = _extrair_celula(row, idx_nome)
            ra_val = _extrair_celula(row, idx_ra)
            ra = _coagir_ra(ra_val)
            nome = str(nome_val).strip() if nome_val is not None else ""

            if not ra:
                continue

            if not _parece_ra(ra):
                raise LeituraPlanilhaError(
                    f"Aba '{nome_aba}', linha {row_index}: valor de RA invalido ({ra!r})."
                )

            alunos.append(
                {
                    "ra": ra,
                    "nome": nome or f"(sem nome, linha {row_index})",
                }
            )

        if alunos:
            turmas[nome_aba] = alunos

    wb.close()
    return LeituraPlanilha(turmas=turmas, avisos=avisos)


def _extrair_itens_brutos_matriculas(dados: Any) -> list[dict[str, Any]]:
    itens = dados
    if isinstance(itens, dict):
        for chave in ("dados", "matriculas", "items", "data"):
            if isinstance(itens.get(chave), list):
                itens = itens[chave]
                break
    if not isinstance(itens, list):
        return []
    return [item for item in itens if isinstance(item, dict)]


def _heuristica_status_matriculado(dados: Any) -> Optional[int]:
    for item in _extrair_itens_brutos_matriculas(dados):
        id_mat = item.get("id_matricula")
        if id_mat is None:
            continue
        try:
            id_int = int(id_mat)
        except (TypeError, ValueError):
            continue
        if str(item.get("status_matricula_diario", "")).upper() == "MATRICULADO":
            return id_int
    return None


def verificar_aluno(
    cliente: IScholarClient,
    turma: str,
    ra: str,
    nome: str,
) -> ResultadoAluno:
    """Replica a resolucao homologada de aluno + matricula para um RA."""
    resultado = ResultadoAluno(turma=turma, ra=ra, nome=nome, status="")

    try:
        busca = cliente.buscar_aluno(ra=ra)
    except Exception as exc:
        resultado.status = "ERRO_REDE"
        resultado.detalhe = f"Excecao em buscar_aluno: {exc!s}"
        return resultado

    if not busca.sucesso:
        resultado.status = "RA_NAO_ENCONTRADO"
        resultado.detalhe = f"HTTP {busca.status_code}: {busca.mensagem}"
        return resultado

    id_aluno = _extrair_id_aluno_da_resposta(busca.dados)
    if id_aluno is None:
        resultado.status = "RA_NAO_ENCONTRADO"
        resultado.detalhe = "buscar_aluno retornou sucesso mas id_aluno nao foi extraido"
        return resultado

    resultado.id_aluno = id_aluno

    try:
        lista = cliente.listar_matriculas(id_aluno=id_aluno, resolver_id_matricula=True)
    except Exception as exc:
        resultado.status = "ERRO_REDE"
        resultado.detalhe = f"Excecao em listar_matriculas: {exc!s}"
        return resultado

    if lista.sucesso and lista.id_matricula_resolvido is not None:
        resultado.status = "OK"
        resultado.id_matricula = lista.id_matricula_resolvido
        return resultado

    ids_ambiguos: list[int] = (lista.rastreabilidade or {}).get("id_matriculas_extraiados", [])

    if len(ids_ambiguos) == 0:
        for situacao in ("cursando", "CURSANDO"):
            try:
                tentativa = cliente.listar_matriculas(
                    id_aluno=id_aluno,
                    resolver_id_matricula=True,
                    situacao=situacao,
                )
            except Exception:
                continue
            if tentativa.sucesso and tentativa.id_matricula_resolvido is not None:
                resultado.status = "HEURISTICA"
                resultado.id_matricula = tentativa.id_matricula_resolvido
                resultado.heuristica_usada = (
                    f"situacao={situacao} (0 resultados sem filtro)"
                )
                return resultado

        resultado.status = "MATRICULA_INACESSIVEL"
        resultado.detalhe = "0 matriculas retornadas (sem filtro e com filtro CURSANDO)"
        return resultado

    if len(ids_ambiguos) > 1:
        for situacao in ("cursando", "CURSANDO"):
            try:
                tentativa = cliente.listar_matriculas(
                    id_aluno=id_aluno,
                    resolver_id_matricula=True,
                    situacao=situacao,
                )
            except Exception:
                continue
            if tentativa.sucesso and tentativa.id_matricula_resolvido is not None:
                resultado.status = "HEURISTICA"
                resultado.id_matricula = tentativa.id_matricula_resolvido
                resultado.heuristica_usada = (
                    f"situacao={situacao} (desambiguacao de {len(ids_ambiguos)} matriculas)"
                )
                return resultado

        id_matriculado = _heuristica_status_matriculado(lista.dados)
        if id_matriculado is not None:
            resultado.status = "HEURISTICA"
            resultado.id_matricula = id_matriculado
            resultado.heuristica_usada = (
                f"status_matricula_diario=MATRICULADO (entre {len(ids_ambiguos)} matriculas)"
            )
            return resultado

        resultado.status = "MATRICULA_AMBIGUA"
        resultado.detalhe = f"{len(ids_ambiguos)} matriculas distintas: {ids_ambiguos}"
        return resultado

    resultado.status = "MATRICULA_INACESSIVEL"
    resultado.detalhe = lista.mensagem or "matricula nao resolvida"
    return resultado


def imprimir_relatorio(relatorios: list[RelatorioTurma], arquivo=None) -> None:
    out = arquivo or sys.stdout

    def p(msg: str = "") -> None:
        print(msg, file=out)

    sep = "=" * 72
    p(sep)
    p("  VERIFICACAO DE CADASTRO - iScholar")
    p(sep)
    p()

    total_geral = 0
    ok_geral = 0
    heur_geral = 0
    falha_geral = 0

    for rel in relatorios:
        total_geral += rel.total
        ok_geral += rel.ok
        heur_geral += rel.heuristica
        falha_geral += rel.falhas

        if rel.fontes:
            p(f"  TURMA: {rel.turma}  ({rel.total} alunos)  [abas: {', '.join(rel.fontes)}]")
        else:
            p(f"  TURMA: {rel.turma}  ({rel.total} alunos)")
        p(f"  OK: {rel.ok}   Heuristica: {rel.heuristica}   Falhas: {rel.falhas}")
        p("-" * 72)

        falhas = [a for a in rel.alunos if a.status not in ("OK", "HEURISTICA")]
        if falhas:
            p("  FALHAS:")
            for a in falhas:
                p(f"    RA {a.ra:>5}  {a.nome}")
                p(f"       Status: {a.status}")
                if a.detalhe:
                    p(f"       Detalhe: {a.detalhe}")
            p()

        heurs = [a for a in rel.alunos if a.status == "HEURISTICA"]
        if heurs:
            p("  HEURISTICAS (funcionais, mas sinalizadas):")
            for a in heurs:
                p(f"    RA {a.ra:>5}  {a.nome}")
                p(f"       id_aluno={a.id_aluno}  id_matricula={a.id_matricula}")
                p(f"       Heuristica: {a.heuristica_usada}")
            p()

    p(sep)
    p("  RESUMO GERAL")
    p(sep)
    p(f"  Total de alunos verificados: {total_geral}")
    p(f"  OK:          {ok_geral}")
    p(f"  Heuristica:  {heur_geral}")
    p(f"  Falhas:      {falha_geral}")
    p()


def salvar_csv(relatorios: list[RelatorioTurma], caminho: str) -> None:
    with open(caminho, "w", newline="", encoding="utf-8-sig") as arquivo:
        writer = csv.writer(arquivo)
        writer.writerow(
            [
                "Turma",
                "RA",
                "Nome",
                "Status",
                "id_aluno",
                "id_matricula",
                "Detalhe",
                "Heuristica",
            ]
        )
        for rel in relatorios:
            for aluno in rel.alunos:
                writer.writerow(
                    [
                        aluno.turma,
                        aluno.ra,
                        aluno.nome,
                        aluno.status,
                        aluno.id_aluno or "",
                        aluno.id_matricula or "",
                        aluno.detalhe,
                        aluno.heuristica_usada,
                    ]
                )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Verifica cadastro de alunos no iScholar a partir de planilha multi-aba.",
    )
    parser.add_argument("planilha", help="Arquivo .xlsx com abas de turma.")
    parser.add_argument("--aba", default=None, help="Verificar apenas uma aba especifica.")
    parser.add_argument("--output", default=None, help="Salvar relatorio em arquivo texto.")
    parser.add_argument("--csv", default=None, dest="csv_output", help="Salvar resultado em CSV.")
    parser.add_argument(
        "--sem-dedup",
        action="store_true",
        default=False,
        dest="sem_dedup",
        help=(
            "Desativa a deduplicacao automatica de alunos por turma+RA no modo Plano B. "
            "Cada aba e verificada individualmente, mesmo que o mesmo aluno apareca "
            "em T1, T2 e T3."
        ),
    )
    args = parser.parse_args()

    if not Path(args.planilha).exists():
        print(f"ERRO: Arquivo nao encontrado: {args.planilha}", file=sys.stderr)
        return 1

    token = os.getenv("ISCHOLAR_API_TOKEN")
    codigo_escola = os.getenv("ISCHOLAR_CODIGO_ESCOLA")
    if not token or not codigo_escola:
        print(
            "ERRO: ISCHOLAR_API_TOKEN e ISCHOLAR_CODIGO_ESCOLA precisam estar definidos.",
            file=sys.stderr,
        )
        return 1

    print(f"Lendo planilha: {args.planilha}")
    try:
        leitura = ler_planilha_multi_aba(args.planilha, aba_filtro=args.aba)
    except LeituraPlanilhaError as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1

    if not leitura.turmas:
        print("AVISO: Nenhuma aba com alunos encontrada.", file=sys.stderr)
        return 1

    for aviso in leitura.avisos:
        print(f"AVISO: {aviso}")

    # Detecção e aplicação de deduplicação Plano B
    tem_abas_plano_b = any(
        parsear_nome_aba_plano_b(nome) is not None for nome in leitura.turmas
    )
    fontes_turma: dict[str, list[str]] = {}

    if tem_abas_plano_b and not args.aba and not args.sem_dedup:
        print("Modo anual detectado — deduplicando alunos por turma+RA (use --sem-dedup para desativar).")
        turmas_proc, fontes_turma, avisos_dedup = deduplicar_turmas_plano_b(leitura.turmas)
        for av in avisos_dedup:
            print(f"  INFO: {av}")
    else:
        turmas_proc = leitura.turmas
        if args.aba:
            print(f"Modo aba unica: {args.aba}")
        elif args.sem_dedup and tem_abas_plano_b:
            print("Deduplicacao desativada (--sem-dedup): cada aba sera verificada individualmente.")

    for nome_chave, alunos in turmas_proc.items():
        label = (
            f"Turma '{nome_chave}' ({', '.join(fontes_turma[nome_chave])})"
            if nome_chave in fontes_turma
            else f"Aba '{nome_chave}'"
        )
        print(f"  {label}: {len(alunos)} alunos")

    total_alunos = sum(len(alunos) for alunos in turmas_proc.values())
    print(f"Total: {total_alunos} alunos em {len(turmas_proc)} turma(s)")
    print()

    cliente = IScholarClient(x_autorizacao=token, x_codigo_escola=codigo_escola)

    relatorios: list[RelatorioTurma] = []
    contador = 0
    for nome_chave, alunos in sorted(turmas_proc.items()):
        rel = RelatorioTurma(
            turma=nome_chave,
            total=len(alunos),
            fontes=fontes_turma.get(nome_chave, []),
        )
        for aluno in alunos:
            contador += 1
            print(
                f"  [{contador}/{total_alunos}] {nome_chave} - RA {aluno['ra']} {aluno['nome'][:40]}...",
                end=" ",
                flush=True,
            )
            resultado = verificar_aluno(cliente, nome_chave, aluno["ra"], aluno["nome"])
            if resultado.status == "OK":
                rel.ok += 1
                print("OK")
            elif resultado.status == "HEURISTICA":
                rel.heuristica += 1
                print(f"HEURISTICA ({resultado.heuristica_usada})")
            else:
                rel.falhas += 1
                print(resultado.status)

            rel.alunos.append(resultado)
            time.sleep(0.2)

        relatorios.append(rel)
        print()

    if args.output:
        with open(args.output, "w", encoding="utf-8") as arquivo:
            imprimir_relatorio(relatorios, arquivo=arquivo)
        print(f"Relatorio salvo em: {args.output}")
        imprimir_relatorio(relatorios)
    else:
        imprimir_relatorio(relatorios)

    if args.csv_output:
        salvar_csv(relatorios, args.csv_output)
        print(f"CSV salvo em: {args.csv_output}")

    total_falhas = sum(rel.falhas for rel in relatorios)
    return 1 if total_falhas > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
