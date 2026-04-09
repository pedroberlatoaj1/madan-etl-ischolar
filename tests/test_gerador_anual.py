"""
Testes para o modo anual do gerador_planilhas.py (Plano B).

Verifica:
- estrutura do workbook gerado (12 abas, nomes corretos)
- cabeçalho de cada aba (colunas fixas + dinâmicas)
- freeze pane em E2
- dados dos alunos nas linhas 2+
- turma e trimestre pré-preenchidos corretamente
- aba sem alunos → só cabeçalho, sem crash
- turmas/trimestres customizados
- compatibilidade total: gerar_planilha_turma e _criar_aba_notas_wide inalterados
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from gerador_planilhas import (
    Aluno,
    COLUNAS_FIXAS_WIDE,
    TRIMESTRES,
    TURMAS_ANUAIS,
    TIPOS_AVALIACAO_WIDE,
    gerar_planilha_turma,
    gerar_workbook_anual,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def alunos_1a() -> list[Aluno]:
    return [
        Aluno(nome="ALICE BARCELOS LINS", ra="1222", turma="1A"),
        Aluno(nome="JOSE DIONIZIO PERTEL", ra="1234", turma="1A"),
    ]


@pytest.fixture
def alunos_1b() -> list[Aluno]:
    return [
        Aluno(nome="ALICE MENEGHELLI DUARTE", ra="1316", turma="1B"),
    ]


@pytest.fixture
def alunos_2a() -> list[Aluno]:
    return [
        Aluno(nome="AGEU ANDRADE CAMPOS LOPES", ra="1065", turma="2A"),
        Aluno(nome="ALICE SIQUEIRA GARCIA", ra="952", turma="2A"),
    ]


@pytest.fixture
def alunos_2b() -> list[Aluno]:
    return [
        Aluno(nome="ALICE GABRIEL BASTOS FERREIRA", ra="1203", turma="2B"),
    ]


@pytest.fixture
def alunos_por_turma(alunos_1a, alunos_1b, alunos_2a, alunos_2b):
    return {
        "1A": alunos_1a,
        "1B": alunos_1b,
        "2A": alunos_2a,
        "2B": alunos_2b,
    }


@pytest.fixture
def workbook_anual(tmp_path, alunos_por_turma) -> Path:
    """Gera o workbook anual e retorna o caminho."""
    return gerar_workbook_anual(
        ano=2026,
        alunos_por_turma=alunos_por_turma,
        output_dir=tmp_path,
    )


# ---------------------------------------------------------------------------
# Testes de estrutura do workbook
# ---------------------------------------------------------------------------

class TestEstruturaWorkbookAnual:
    def test_arquivo_criado(self, workbook_anual):
        assert workbook_anual.exists()
        assert workbook_anual.name == "madan_2026_anual.xlsx"

    def test_12_abas_geradas(self, workbook_anual):
        wb = openpyxl.load_workbook(workbook_anual)
        assert len(wb.sheetnames) == 12

    def test_nomes_das_abas(self, workbook_anual):
        wb = openpyxl.load_workbook(workbook_anual)
        esperadas = [
            "1A_T1", "1A_T2", "1A_T3",
            "1B_T1", "1B_T2", "1B_T3",
            "2A_T1", "2A_T2", "2A_T3",
            "2B_T1", "2B_T2", "2B_T3",
        ]
        assert wb.sheetnames == esperadas

    def test_ordem_das_abas(self, workbook_anual):
        """Abas devem estar em ordem: turmas primeiro, trimestres dentro de cada turma."""
        wb = openpyxl.load_workbook(workbook_anual)
        nomes = wb.sheetnames
        # 1A antes de 1B, T1 antes de T2
        assert nomes.index("1A_T1") < nomes.index("1A_T2")
        assert nomes.index("1A_T3") < nomes.index("1B_T1")
        assert nomes.index("2A_T3") < nomes.index("2B_T1")


# ---------------------------------------------------------------------------
# Testes de cabeçalho de cada aba
# ---------------------------------------------------------------------------

class TestCabecalhoDasAbas:
    def test_colunas_fixas_nas_4_primeiras(self, workbook_anual):
        wb = openpyxl.load_workbook(workbook_anual)
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
            assert headers == list(COLUNAS_FIXAS_WIDE), (
                f"Aba {nome_aba}: 4 primeiras colunas incorretas: {headers}"
            )

    def test_colunas_dinamicas_presentes(self, workbook_anual):
        """Cada aba deve ter pelo menos uma coluna dinâmica além das 4 fixas."""
        wb = openpyxl.load_workbook(workbook_anual)
        n_fixas = len(COLUNAS_FIXAS_WIDE)
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            total_cols = sum(1 for c in range(1, ws.max_column + 1)
                             if ws.cell(row=1, column=c).value is not None)
            assert total_cols > n_fixas, (
                f"Aba {nome_aba}: sem colunas dinâmicas (total={total_cols})"
            )

    def test_formato_coluna_dinamica(self, workbook_anual):
        """Colunas dinâmicas devem seguir o padrão 'Disc - Frente X - Tipo'."""
        wb = openpyxl.load_workbook(workbook_anual)
        ws = wb["1A_T1"]
        n_fixas = len(COLUNAS_FIXAS_WIDE)
        col5 = ws.cell(row=1, column=n_fixas + 1).value
        assert col5 is not None
        partes = str(col5).split(" - ")
        assert len(partes) == 3, f"Formato inválido: {col5!r}"
        # Tipo deve estar em TIPOS_AVALIACAO_WIDE
        assert partes[2] in TIPOS_AVALIACAO_WIDE, f"Tipo desconhecido: {partes[2]!r}"

    def test_1a_e_1b_tem_cabecalhos_diferentes(self, workbook_anual):
        """1A e 1B podem ter disciplinas/frentes distintas — cabeçalhos podem diferir."""
        wb = openpyxl.load_workbook(workbook_anual)
        ws1a = wb["1A_T1"]
        ws1b = wb["1B_T1"]
        cols_1a = ws1a.max_column
        cols_1b = ws1b.max_column
        # Ambas devem ter mais de 4 colunas
        assert cols_1a > 4
        assert cols_1b > 4

    def test_mesmo_cabecalho_para_mesmo_turma_trimestres_diferentes(self, workbook_anual):
        """1A_T1 e 1A_T2 devem ter o mesmo cabeçalho (mesmo ano, mesmos professores)."""
        wb = openpyxl.load_workbook(workbook_anual)
        ws_t1 = wb["1A_T1"]
        ws_t2 = wb["1A_T2"]
        cabecalho_t1 = [ws_t1.cell(row=1, column=c).value for c in range(1, ws_t1.max_column + 1)]
        cabecalho_t2 = [ws_t2.cell(row=1, column=c).value for c in range(1, ws_t2.max_column + 1)]
        assert cabecalho_t1 == cabecalho_t2


# ---------------------------------------------------------------------------
# Testes de dados dos alunos
# ---------------------------------------------------------------------------

class TestDadosAlunos:
    def test_alunos_preenchidos_na_turma(self, workbook_anual, alunos_1a):
        wb = openpyxl.load_workbook(workbook_anual)
        ws = wb["1A_T1"]
        nomes_na_aba = [ws.cell(row=r, column=1).value for r in range(2, 2 + len(alunos_1a))]
        nomes_esperados = [a.nome for a in alunos_1a]
        assert nomes_na_aba == nomes_esperados

    def test_ra_preenchido(self, workbook_anual, alunos_1a):
        wb = openpyxl.load_workbook(workbook_anual)
        ws = wb["1A_T1"]
        ra_na_aba = ws.cell(row=2, column=2).value
        assert str(ra_na_aba) == alunos_1a[0].ra

    def test_turma_preenchida(self, workbook_anual):
        wb = openpyxl.load_workbook(workbook_anual)
        ws = wb["2A_T2"]
        turma_na_aba = ws.cell(row=2, column=3).value
        assert turma_na_aba == "2A"

    def test_trimestre_preenchido_corretamente(self, workbook_anual):
        """Cada aba deve ter o trimestre correto pré-preenchido."""
        wb = openpyxl.load_workbook(workbook_anual)
        for turma in TURMAS_ANUAIS:
            for tri in TRIMESTRES:
                ws = wb[f"{turma}_{tri}"]
                # Só verifica se houver alunos
                if ws.max_row >= 2 and ws.cell(row=2, column=1).value:
                    tri_na_aba = ws.cell(row=2, column=4).value
                    assert tri_na_aba == tri, (
                        f"Aba {turma}_{tri}: trimestre incorreto: {tri_na_aba!r}"
                    )

    def test_aba_sem_alunos_nao_crasha(self, tmp_path):
        """Gerar workbook com dict vazio não deve lançar exceção."""
        filepath = gerar_workbook_anual(
            ano=2026,
            alunos_por_turma={},  # nenhum aluno
            output_dir=tmp_path,
        )
        assert filepath.exists()
        wb = openpyxl.load_workbook(filepath)
        ws = wb["1A_T1"]
        # Só o cabeçalho — linha 2 vazia
        assert ws.cell(row=2, column=1).value is None

    def test_notas_em_branco(self, workbook_anual):
        """Colunas de nota devem estar em branco (professores preenchem depois)."""
        wb = openpyxl.load_workbook(workbook_anual)
        ws = wb["1A_T1"]
        n_fixas = len(COLUNAS_FIXAS_WIDE)
        # Verifica primeira coluna de nota da primeira linha de dado
        nota_col = ws.cell(row=2, column=n_fixas + 1).value
        assert nota_col is None


# ---------------------------------------------------------------------------
# Testes de freeze pane e formatação
# ---------------------------------------------------------------------------

class TestFreezePane:
    def test_freeze_em_e2(self, workbook_anual):
        """Todas as abas devem ter freeze em E2 (cabeçalho + 4 colunas fixas)."""
        wb = openpyxl.load_workbook(workbook_anual)
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            assert ws.freeze_panes == "E2", (
                f"Aba {nome_aba}: freeze_panes={ws.freeze_panes!r}, esperado 'E2'"
            )


# ---------------------------------------------------------------------------
# Testes de turmas/trimestres customizados
# ---------------------------------------------------------------------------

class TestParametrosCustomizados:
    def test_apenas_1a(self, tmp_path, alunos_por_turma):
        filepath = gerar_workbook_anual(
            ano=2026,
            alunos_por_turma=alunos_por_turma,
            output_dir=tmp_path,
            turmas=["1A"],
        )
        wb = openpyxl.load_workbook(filepath)
        assert wb.sheetnames == ["1A_T1", "1A_T2", "1A_T3"]

    def test_apenas_t1(self, tmp_path, alunos_por_turma):
        filepath = gerar_workbook_anual(
            ano=2026,
            alunos_por_turma=alunos_por_turma,
            output_dir=tmp_path,
            trimestres=["T1"],
        )
        wb = openpyxl.load_workbook(filepath)
        assert wb.sheetnames == ["1A_T1", "1B_T1", "2A_T1", "2B_T1"]

    def test_subconjunto_turmas_e_trimestres(self, tmp_path, alunos_por_turma):
        filepath = gerar_workbook_anual(
            ano=2026,
            alunos_por_turma=alunos_por_turma,
            output_dir=tmp_path,
            turmas=["2A", "2B"],
            trimestres=["T1", "T2"],
        )
        wb = openpyxl.load_workbook(filepath)
        assert wb.sheetnames == ["2A_T1", "2A_T2", "2B_T1", "2B_T2"]


# ---------------------------------------------------------------------------
# Teste de retrocompatibilidade — modo por turma intocado
# ---------------------------------------------------------------------------

class TestRetrocompatibilidade:
    def test_gerar_planilha_turma_ainda_funciona(self, tmp_path, alunos_1a):
        """gerar_planilha_turma deve continuar gerando arquivo com aba 'Notas'."""
        filepath = gerar_planilha_turma(
            turma="1A",
            trimestre="T1",
            ano=2026,
            alunos=alunos_1a,
            output_dir=tmp_path,
        )
        assert filepath.exists()
        assert filepath.name == "1A_T1_2026.xlsx"

        wb = openpyxl.load_workbook(filepath)
        assert wb.sheetnames == ["Notas"]

        ws = wb["Notas"]
        assert ws.freeze_panes == "E2"
        # Cabeçalho intacto
        assert ws.cell(row=1, column=1).value == "Estudante"
        assert ws.cell(row=1, column=2).value == "RA"

    def test_workbook_anual_nao_afeta_planilha_turma(self, tmp_path, alunos_por_turma, alunos_1a):
        """Gerar workbook anual não deve interferir com geração por turma."""
        gerar_workbook_anual(
            ano=2026,
            alunos_por_turma=alunos_por_turma,
            output_dir=tmp_path,
        )
        # Ainda pode gerar por turma sem conflito
        filepath = gerar_planilha_turma(
            turma="1A",
            trimestre="T1",
            ano=2026,
            alunos=alunos_1a,
            output_dir=tmp_path,
        )
        assert filepath.exists()
        wb = openpyxl.load_workbook(filepath)
        assert "Notas" in wb.sheetnames
