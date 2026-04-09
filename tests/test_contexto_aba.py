"""
Testes para parsear_nome_aba() e aplicar_contexto_aba() — derivação de
Turma/Trimestre a partir do nome da aba no workbook anual (Plano B).

Cobre:
- parsear_nome_aba: padrões válidos e inválidos
- aplicar_contexto_aba: campos ausentes, vazios, coerentes, conflitantes
- retrocompatibilidade: nome de aba legado não altera o df
- integração com carregar_entrada: injeção transparente
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from pipeline_runner import (
    TemplateInvalidoError,
    aplicar_contexto_aba,
    carregar_entrada,
    parsear_nome_aba,
)


# ---------------------------------------------------------------------------
# Testes de parsear_nome_aba
# ---------------------------------------------------------------------------

class TestParsearNomeAba:
    @pytest.mark.parametrize("nome,esperado", [
        ("2A_T1", ("2A", "T1")),
        ("1B_T3", ("1B", "T3")),
        ("2B_T2", ("2B", "T2")),
        ("1A_T1", ("1A", "T1")),
        # Case-insensitive
        ("2a_t1", ("2A", "T1")),
        ("1B_t2", ("1B", "T2")),
    ])
    def test_padroes_validos(self, nome, esperado):
        assert parsear_nome_aba(nome) == esperado

    @pytest.mark.parametrize("nome", [
        "Notas",        # aba legada
        "Sheet1",       # Excel default
        "1A",           # sem trimestre
        "T1",           # sem turma
        "2A_T4",        # T4 não existe
        "2A_T0",        # T0 não existe
        "22A_T1",       # série de 2 dígitos
        "2A-T1",        # separador errado
        "2A T1",        # espaço
        "",             # vazio
        "  ",           # só espaços
    ])
    def test_padroes_invalidos_retornam_none(self, nome):
        assert parsear_nome_aba(nome) is None

    def test_none_retorna_none(self):
        assert parsear_nome_aba(None) is None  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Testes de aplicar_contexto_aba — campos ausentes
# ---------------------------------------------------------------------------

class TestAplicarContextoAba:

    def _df_sem_turma_trimestre(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Arte - Frente Única - AV 1 Obj": "7"},
            {"Estudante": "JOSE",  "RA": "1234", "Arte - Frente Única - AV 1 Obj": "8"},
        ])

    def _df_com_turma_trimestre_vazios(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "", "Trimestre": ""},
            {"Estudante": "JOSE",  "RA": "1234", "Turma": None, "Trimestre": None},
        ])

    def _df_com_turma_trimestre_corretos(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "2A", "Trimestre": "T1"},
            {"Estudante": "JOSE",  "RA": "1234", "Turma": "2A", "Trimestre": "T1"},
        ])

    # --- Campos ausentes ---

    def test_colunas_ausentes_sao_criadas(self):
        df = self._df_sem_turma_trimestre()
        resultado = aplicar_contexto_aba(df, "2A_T1")
        assert "Turma" in resultado.columns
        assert "Trimestre" in resultado.columns
        assert (resultado["Turma"] == "2A").all()
        assert (resultado["Trimestre"] == "T1").all()

    def test_colunas_ausentes_valores_corretos_por_turma(self):
        df = self._df_sem_turma_trimestre()
        resultado = aplicar_contexto_aba(df, "1B_T3")
        assert (resultado["Turma"] == "1B").all()
        assert (resultado["Trimestre"] == "T3").all()

    # --- Campos vazios ---

    def test_campos_vazios_sao_preenchidos(self):
        df = self._df_com_turma_trimestre_vazios()
        resultado = aplicar_contexto_aba(df, "2A_T1")
        assert (resultado["Turma"] == "2A").all()
        assert (resultado["Trimestre"] == "T1").all()

    def test_campo_none_e_preenchido(self):
        df = pd.DataFrame([{"Estudante": "X", "RA": "1", "Turma": None, "Trimestre": None}])
        resultado = aplicar_contexto_aba(df, "1A_T2")
        assert resultado.iloc[0]["Turma"] == "1A"
        assert resultado.iloc[0]["Trimestre"] == "T2"

    # --- Campos já preenchidos e coerentes ---

    def test_campos_coerentes_nao_sao_sobrescritos(self):
        df = self._df_com_turma_trimestre_corretos()
        resultado = aplicar_contexto_aba(df, "2A_T1")
        # Valores originais preservados (não reescritos desnecessariamente)
        assert (resultado["Turma"] == "2A").all()
        assert (resultado["Trimestre"] == "T1").all()

    def test_campos_coerentes_case_insensitive(self):
        """Aba '2a_t1' deve ser coerente com Turma='2A', Trimestre='T1'."""
        df = self._df_com_turma_trimestre_corretos()
        resultado = aplicar_contexto_aba(df, "2a_t1")
        assert (resultado["Turma"] == "2A").all()

    # --- Conflito ---

    def test_conflito_turma_lanca_erro(self):
        df = pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "1A", "Trimestre": "T1"},
        ])
        with pytest.raises(TemplateInvalidoError) as exc_info:
            aplicar_contexto_aba(df, "2A_T1")
        msg = str(exc_info.value)
        assert "2A_T1" in msg
        assert "Turma" in msg
        assert "1A" in msg
        assert "2A" in msg

    def test_conflito_trimestre_lanca_erro(self):
        df = pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "2A", "Trimestre": "T3"},
        ])
        with pytest.raises(TemplateInvalidoError) as exc_info:
            aplicar_contexto_aba(df, "2A_T1")
        msg = str(exc_info.value)
        assert "Trimestre" in msg
        assert "T3" in msg
        assert "T1" in msg

    def test_conflito_em_linha_especifica_menciona_linha(self):
        """A mensagem de erro deve indicar a linha exata com conflito."""
        df = pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "2A", "Trimestre": "T1"},
            {"Estudante": "JOSE",  "RA": "1234", "Turma": "1B", "Trimestre": "T1"},  # conflito
        ])
        with pytest.raises(TemplateInvalidoError) as exc_info:
            aplicar_contexto_aba(df, "2A_T1")
        msg = str(exc_info.value)
        assert "linha 3" in msg  # linha 2 = idx 0 (ok), linha 3 = idx 1 (conflito)

    def test_conflito_parcial_linha_mista(self):
        """Uma linha vazia + uma com conflito → deve falhar."""
        df = pd.DataFrame([
            {"Estudante": "ALICE", "RA": "1222", "Turma": "",   "Trimestre": "T1"},  # turma vazia
            {"Estudante": "JOSE",  "RA": "1234", "Turma": "1A", "Trimestre": "T1"},  # conflito
        ])
        with pytest.raises(TemplateInvalidoError):
            aplicar_contexto_aba(df, "2A_T1")

    # --- df original não é modificado in-place ---

    def test_df_original_nao_modificado(self):
        df = self._df_sem_turma_trimestre()
        cols_antes = set(df.columns)
        aplicar_contexto_aba(df, "2A_T1")
        assert set(df.columns) == cols_antes  # df original intocado

    # --- Retrocompatibilidade ---

    def test_nome_aba_legado_nao_altera_df(self):
        """Aba com nome legado (ex: 'Notas') não deve alterar o df."""
        df = self._df_com_turma_trimestre_corretos()
        resultado = aplicar_contexto_aba(df, "Notas")
        pd.testing.assert_frame_equal(df, resultado)

    def test_nome_aba_none_nao_altera_df(self):
        df = self._df_com_turma_trimestre_corretos()
        resultado = aplicar_contexto_aba(df, None)
        pd.testing.assert_frame_equal(df, resultado)

    def test_df_vazio_nao_crasha(self):
        df = pd.DataFrame(columns=["Estudante", "RA"])
        resultado = aplicar_contexto_aba(df, "2A_T1")
        assert "Turma" in resultado.columns
        assert "Trimestre" in resultado.columns
        assert len(resultado) == 0


# ---------------------------------------------------------------------------
# Integração: carregar_entrada + aplicar_contexto_aba
# ---------------------------------------------------------------------------

class TestIntegracaoCarregarEntrada:
    """
    Verifica que carregar_entrada injeta Turma/Trimestre automaticamente
    quando sheet_name segue o padrão Plano B.
    """

    @pytest.fixture
    def xlsx_multi_aba_sem_contexto(self, tmp_path) -> Path:
        """Workbook anual cujas abas NÃO têm Turma/Trimestre pré-preenchidos."""
        wb = openpyxl.Workbook()
        for nome in ["1A_T1", "2A_T2"]:
            ws = wb.create_sheet(nome)
            ws.append(["Estudante", "RA", "Arte - Frente Única - AV 1 Obj"])
            ws.append(["ALICE", "1222", "7"])
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        caminho = tmp_path / "anual.xlsx"
        wb.save(caminho)
        return caminho

    @pytest.fixture
    def xlsx_aba_unica_com_contexto(self, tmp_path) -> Path:
        """Aba única (legado) com Turma/Trimestre já preenchidos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notas"
        ws.append(["Estudante", "RA", "Turma", "Trimestre", "Arte - Frente Única - AV 1 Obj"])
        ws.append(["ALICE", "1222", "1A", "T1", "7"])
        caminho = tmp_path / "legado.xlsx"
        wb.save(caminho)
        return caminho

    def test_aba_anual_injeta_turma_trimestre(self, xlsx_multi_aba_sem_contexto):
        df = carregar_entrada(xlsx_multi_aba_sem_contexto, sheet_name="2A_T2")
        assert df.iloc[0]["Turma"] == "2A"
        assert df.iloc[0]["Trimestre"] == "T2"

    def test_aba_anual_turma_correta_por_aba(self, xlsx_multi_aba_sem_contexto):
        df = carregar_entrada(xlsx_multi_aba_sem_contexto, sheet_name="1A_T1")
        assert df.iloc[0]["Turma"] == "1A"
        assert df.iloc[0]["Trimestre"] == "T1"

    def test_aba_legada_preserva_contexto_existente(self, xlsx_aba_unica_com_contexto):
        """Planilha legada (aba 'Notas') não deve ter Turma/Trimestre alterados."""
        df = carregar_entrada(xlsx_aba_unica_com_contexto)
        assert df.iloc[0]["Turma"] == "1A"
        assert df.iloc[0]["Trimestre"] == "T1"

    def test_conflito_dentro_do_carregar_entrada_lanca_template_invalido(self, tmp_path):
        """Se a aba diz 2A_T1 mas as células dizem Turma=1B, deve falhar em carregar_entrada."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("2A_T1")
        ws.append(["Estudante", "RA", "Turma", "Trimestre"])
        ws.append(["ALICE", "1222", "1B", "T1"])  # Turma errada
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        caminho = tmp_path / "conflito.xlsx"
        wb.save(caminho)

        with pytest.raises(TemplateInvalidoError) as exc_info:
            carregar_entrada(caminho, sheet_name="2A_T1")
        msg = str(exc_info.value)
        assert "2A_T1" in msg
        assert "1B" in msg
        assert "2A" in msg
