"""
Testes para compilador_turma.py — compilação de planilha multi-abas para formato pipeline.

NOTA: O compilador_turma.py opera no formato antigo (multi-abas). Os fixtures
criam planilhas nesse formato diretamente via openpyxl, sem depender do
gerar_planilhas.py (que agora gera formato wide).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from compilador_turma import (
    _formatar_frente_professor,
    _ler_metadata,
    compilar_planilha_turma,
    compilar_planilha_para_arquivo,
    compilar_diretorio,
)
from gerador_planilhas import Aluno, COLUNAS_NOTA


# ---------------------------------------------------------------------------
# Helper — cria planilha no formato antigo (multi-abas + _metadata)
# ---------------------------------------------------------------------------

COLUNAS_IDENTIDADE_ANTIGO = ["Nome", "RA", "Turma"]

COLUNAS_CONFERENCIA_ANTIGO = ["Nota sem a AV 3", "Nota com a AV 3", "Nota Final"]

TODAS_COLUNAS_ANTIGO = COLUNAS_IDENTIDADE_ANTIGO + COLUNAS_NOTA + COLUNAS_CONFERENCIA_ANTIGO


def _criar_planilha_formato_antigo(
    path: Path,
    turma: str,
    trimestre: str,
    ano: int,
    alunos: list[Aluno],
    abas: list[dict],
) -> Path:
    """
    Cria uma planilha no formato antigo (multi-abas + _metadata) para testes.

    abas: list de dicts com keys 'disciplina', 'frente', 'professor'
    """
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    for aba_info in abas:
        disc = aba_info["disciplina"]
        frente = aba_info.get("frente", "")
        prof = aba_info.get("professor", "Prof")
        if frente:
            nome_aba = f"{disc}_{frente}_{prof}"[:31]
        else:
            nome_aba = f"{disc}_{prof}"[:31]

        ws = wb.create_sheet(title=nome_aba)

        # Header
        for c, col in enumerate(TODAS_COLUNAS_ANTIGO, 1):
            ws.cell(row=1, column=c, value=col)

        # Dados dos alunos
        for r, aluno in enumerate(alunos, 2):
            ws.cell(row=r, column=1, value=aluno.nome)
            ws.cell(row=r, column=2, value=aluno.ra)
            ws.cell(row=r, column=3, value=turma)

    # Aba _metadata
    ws_meta = wb.create_sheet(title="_metadata")
    ws_meta.cell(row=1, column=1, value="chave")
    ws_meta.cell(row=1, column=2, value="valor")
    metadata = [
        ("trimestre", trimestre),
        ("turma", turma),
        ("serie", turma[0] if turma[0].isdigit() else ""),
        ("ano", str(ano)),
        ("gerado_em", "2026-03-31"),
        ("total_abas", str(len(abas))),
    ]
    for i, (k, v) in enumerate(metadata, 2):
        ws_meta.cell(row=i, column=1, value=k)
        ws_meta.cell(row=i, column=2, value=v)

    # Mapa de abas
    ws_meta.cell(row=10, column=1, value="nome_aba")
    ws_meta.cell(row=10, column=2, value="disciplina")
    ws_meta.cell(row=10, column=3, value="frente")
    ws_meta.cell(row=10, column=4, value="professor")
    ws_meta.cell(row=10, column=5, value="professor_nome_completo")

    for i, aba_info in enumerate(abas, 11):
        disc = aba_info["disciplina"]
        frente = aba_info.get("frente", "")
        prof = aba_info.get("professor", "Prof")
        if frente:
            nome_aba = f"{disc}_{frente}_{prof}"[:31]
        else:
            nome_aba = f"{disc}_{prof}"[:31]
        ws_meta.cell(row=i, column=1, value=nome_aba)
        ws_meta.cell(row=i, column=2, value=disc)
        ws_meta.cell(row=i, column=3, value=frente)
        ws_meta.cell(row=i, column=4, value=prof)
        ws_meta.cell(row=i, column=5, value=prof)

    ws_meta.sheet_state = "hidden"

    wb.save(str(path))
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

ABAS_PADRAO = [
    {"disciplina": "matematica", "frente": "F1", "professor": "Luan"},
    {"disciplina": "fisica", "frente": "F1", "professor": "Cavaco"},
]


@pytest.fixture
def alunos_1a() -> list[Aluno]:
    return [
        Aluno(nome="Alice Silva", ra="1001", turma="1A"),
        Aluno(nome="Bruno Costa", ra="1002", turma="1A"),
    ]


@pytest.fixture
def planilha_1a(alunos_1a: list[Aluno], tmp_path: Path) -> Path:
    """Gera planilha no formato antigo (multi-abas) para turma 1A."""
    path = tmp_path / "1A_T1_2026.xlsx"
    return _criar_planilha_formato_antigo(
        path, "1A", "T1", 2026, alunos_1a, ABAS_PADRAO,
    )


@pytest.fixture
def planilha_1a_com_notas(planilha_1a: Path) -> Path:
    """Planilha 1A com notas preenchidas na primeira aba."""
    wb = openpyxl.load_workbook(str(planilha_1a))

    abas_disc = [s for s in wb.sheetnames if s != "_metadata"]
    ws = wb[abas_disc[0]]

    # Alice (row 2): AV1(OBJ)=4, AV1(DISC)=3
    ws.cell(row=2, column=4, value=4)   # AV 1 (OBJ)
    ws.cell(row=2, column=5, value=3)   # AV 1 (DISC)

    # Bruno (row 3): AV1(OBJ)=5, AV1(DISC)=4, Simulado=7
    ws.cell(row=3, column=4, value=5)
    ws.cell(row=3, column=5, value=4)
    ws.cell(row=3, column=10, value=7)  # Simulado

    wb.save(str(planilha_1a))
    wb.close()
    return planilha_1a


# ---------------------------------------------------------------------------
# TestFormatarFrenteProfessor
# ---------------------------------------------------------------------------

class TestFormatarFrenteProfessor:
    def test_com_frente(self):
        assert _formatar_frente_professor("F2", "Luan") == "F2 - Luan"

    def test_sem_frente(self):
        assert _formatar_frente_professor("", "Nery") == "Nery"

    def test_frente_complexa(self):
        assert _formatar_frente_professor("F1", "Cavaco") == "F1 - Cavaco"


# ---------------------------------------------------------------------------
# TestLerMetadata
# ---------------------------------------------------------------------------

class TestLerMetadata:
    def test_le_metadata_corretamente(self, planilha_1a: Path):
        wb = openpyxl.load_workbook(str(planilha_1a))
        meta = _ler_metadata(wb)
        wb.close()

        assert meta["trimestre"] == "T1"
        assert meta["turma"] == "1A"
        assert meta["ano"] == "2026"
        assert len(meta["mapa_abas"]) > 0

    def test_mapa_abas_tem_disciplina(self, planilha_1a: Path):
        wb = openpyxl.load_workbook(str(planilha_1a))
        meta = _ler_metadata(wb)
        wb.close()

        for aba in meta["mapa_abas"]:
            assert aba["disciplina"], f"Aba {aba['nome_aba']} sem disciplina"
            assert aba["professor"], f"Aba {aba['nome_aba']} sem professor"

    def test_sem_metadata_levanta_erro(self, tmp_path: Path):
        wb = openpyxl.Workbook()
        wb.active.title = "dados"
        path = tmp_path / "sem_meta.xlsx"
        wb.save(str(path))
        wb.close()

        wb2 = openpyxl.load_workbook(str(path))
        with pytest.raises(ValueError, match="_metadata"):
            _ler_metadata(wb2)
        wb2.close()


# ---------------------------------------------------------------------------
# TestCompilarPlanilha
# ---------------------------------------------------------------------------

class TestCompilarPlanilha:
    def test_planilha_sem_notas_retorna_vazio(self, planilha_1a: Path):
        df = compilar_planilha_turma(planilha_1a)
        assert df.empty

    def test_planilha_com_notas_retorna_linhas(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        assert not df.empty
        assert len(df) == 2  # Alice e Bruno

    def test_colunas_pipeline_presentes(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        colunas_obrigatorias = [
            "Estudante", "RA", "Turma", "Trimestre",
            "Disciplina", "Frente - Professor",
        ]
        for col in colunas_obrigatorias:
            assert col in df.columns, f"Coluna '{col}' ausente no output"

    def test_trimestre_preenchido(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        assert all(df["Trimestre"] == "T1")

    def test_disciplina_preenchida(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        assert all(df["Disciplina"].notna())
        assert all(df["Disciplina"] != "")

    def test_frente_professor_preenchida(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        assert all(df["Frente - Professor"].notna())

    def test_notas_preservadas(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        alice = df[df["Estudante"] == "Alice Silva"].iloc[0]
        assert alice["AV 1 (OBJ)"] == 4
        assert alice["AV 1 (DISC)"] == 3

    def test_ordenado_por_aluno_disciplina(self, planilha_1a_com_notas: Path):
        df = compilar_planilha_turma(planilha_1a_com_notas)
        estudantes = list(df["Estudante"])
        assert estudantes == sorted(estudantes)


# ---------------------------------------------------------------------------
# TestCompilarParaArquivo
# ---------------------------------------------------------------------------

class TestCompilarParaArquivo:
    def test_gera_arquivo_xlsx(self, planilha_1a_com_notas: Path, tmp_path: Path):
        output = tmp_path / "output" / "pipeline.xlsx"
        result = compilar_planilha_para_arquivo(planilha_1a_com_notas, output)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_arquivo_legivel_por_pandas(self, planilha_1a_com_notas: Path, tmp_path: Path):
        output = tmp_path / "pipeline.xlsx"
        compilar_planilha_para_arquivo(planilha_1a_com_notas, output)
        df = pd.read_excel(str(output), dtype=str)
        assert len(df) == 2
        assert "Estudante" in df.columns


# ---------------------------------------------------------------------------
# TestCompilarDiretorio
# ---------------------------------------------------------------------------

class TestCompilarDiretorio:
    def test_compila_todos_xlsx(self, alunos_1a: list[Aluno], tmp_path: Path):
        input_dir = tmp_path / "planilhas"
        input_dir.mkdir()
        output_dir = tmp_path / "pipeline"

        filepath = _criar_planilha_formato_antigo(
            input_dir / "1A_T1_2026.xlsx", "1A", "T1", 2026, alunos_1a, ABAS_PADRAO,
        )

        # Preenche uma nota
        wb = openpyxl.load_workbook(str(filepath))
        abas_disc = [s for s in wb.sheetnames if s != "_metadata"]
        ws = wb[abas_disc[0]]
        ws.cell(row=2, column=4, value=8)  # AV1 OBJ para Alice
        wb.save(str(filepath))
        wb.close()

        arquivos = compilar_diretorio(input_dir, output_dir)
        assert len(arquivos) == 1
        assert "pipeline" in arquivos[0].name


# ---------------------------------------------------------------------------
# TestRoundTrip — teste de integração formato antigo
# ---------------------------------------------------------------------------

class TestRoundTrip:
    """
    Teste round-trip: cria planilha formato antigo → preenche notas → compila
    → verifica que o output tem o formato exato esperado pelo pipeline.
    """

    def test_round_trip_formato_pipeline(self, alunos_1a: list[Aluno], tmp_path: Path):
        filepath = _criar_planilha_formato_antigo(
            tmp_path / "1A_T1_2026.xlsx", "1A", "T1", 2026, alunos_1a, ABAS_PADRAO,
        )

        wb = openpyxl.load_workbook(str(filepath))
        abas_disc = [s for s in wb.sheetnames if s != "_metadata"]
        ws = wb[abas_disc[0]]

        # Alice: AV1(OBJ)=4, AV1(DISC)=3, AV2(OBJ)=5, AV2(DISC)=4
        ws.cell(row=2, column=4, value=4)
        ws.cell(row=2, column=5, value=3)
        ws.cell(row=2, column=6, value=5)
        ws.cell(row=2, column=7, value=4)

        wb.save(str(filepath))
        wb.close()

        df = compilar_planilha_turma(filepath)
        assert len(df) == 1  # Só Alice tem notas

        row = df.iloc[0]
        assert row["Estudante"] == "Alice Silva"
        assert row["RA"] == "1001"
        assert row["Turma"] == "1A"
        assert row["Trimestre"] == "T1"
        assert row["Disciplina"] != ""
        assert row["Frente - Professor"] != ""
        assert row["AV 1 (OBJ)"] == 4
        assert row["AV 1 (DISC)"] == 3
        assert row["AV 2 (OBJ)"] == 5
        assert row["AV 2 (DISC)"] == 4

    def test_round_trip_multiplas_abas(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Verifica que notas em abas diferentes geram linhas distintas."""
        filepath = _criar_planilha_formato_antigo(
            tmp_path / "1A_T1_2026.xlsx", "1A", "T1", 2026, alunos_1a, ABAS_PADRAO,
        )

        wb = openpyxl.load_workbook(str(filepath))
        abas_disc = [s for s in wb.sheetnames if s != "_metadata"]

        # Preenche notas em 2 abas diferentes para Alice
        for aba_name in abas_disc[:2]:
            ws = wb[aba_name]
            ws.cell(row=2, column=4, value=7)  # AV1(OBJ) para Alice

        wb.save(str(filepath))
        wb.close()

        df = compilar_planilha_turma(filepath)

        alice_rows = df[df["Estudante"] == "Alice Silva"]
        assert len(alice_rows) == 2

        disciplinas = list(alice_rows["Disciplina"])
        assert len(set(disciplinas)) == 2

    def test_round_trip_aluno_sem_nota_ignorado(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Alunos sem nenhuma nota preenchida não aparecem no output."""
        filepath = _criar_planilha_formato_antigo(
            tmp_path / "1A_T1_2026.xlsx", "1A", "T1", 2026, alunos_1a, ABAS_PADRAO,
        )

        wb = openpyxl.load_workbook(str(filepath))
        abas_disc = [s for s in wb.sheetnames if s != "_metadata"]
        ws = wb[abas_disc[0]]

        ws.cell(row=2, column=4, value=8)  # Alice: AV1(OBJ)

        wb.save(str(filepath))
        wb.close()

        df = compilar_planilha_turma(filepath)
        assert len(df) == 1
        assert df.iloc[0]["Estudante"] == "Alice Silva"
