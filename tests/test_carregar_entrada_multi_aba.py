"""
Testes para carregar_entrada() e listar_abas_xlsx() — suporte a workbook multi-aba.

Cobre:
- aba única: comportamento original intacto (sheet_name=None)
- multi-aba com sheet_name explícito: lê aba correta
- multi-aba sem sheet_name: TemplateInvalidoError com lista de abas
- aba inexistente: TemplateInvalidoError com nome correto das abas
- DataFrame passado diretamente: ignora sheet_name (sem crash)
- CSV com sheet_name: aviso ignorado, lê normalmente
- listar_abas_xlsx: retorna nomes corretos
- listar_abas_xlsx em CSV: retorna lista vazia
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from pipeline_runner import TemplateInvalidoError, carregar_entrada, listar_abas_xlsx


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def xlsx_aba_unica(tmp_path) -> Path:
    """Xlsx com 1 aba 'Notas', cabeçalho válido."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"
    ws.append(["Estudante", "RA", "Turma", "Trimestre", "Arte - Frente Única - AV 1 Obj"])
    ws.append(["ALICE", "1222", "1A", "T1", "7"])
    caminho = tmp_path / "aba_unica.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def xlsx_multi_aba(tmp_path) -> Path:
    """Xlsx com 4 abas: 1A_T1, 1A_T2, 2A_T1, 2A_T2."""
    wb = openpyxl.Workbook()
    nomes = ["1A_T1", "1A_T2", "2A_T1", "2A_T2"]
    for i, nome in enumerate(nomes):
        ws = wb.create_sheet(title=nome)
        turma = nome.split("_")[0]
        tri = nome.split("_")[1]
        ws.append(["Estudante", "RA", "Turma", "Trimestre", "Arte - Frente Única - AV 1 Obj"])
        ws.append([f"ALUNO_{nome}", f"100{i}", turma, tri, "8"])
    # Remove aba default vazia
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    caminho = tmp_path / "multi_aba.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def csv_simples(tmp_path) -> Path:
    """CSV simples com 1 aluno."""
    caminho = tmp_path / "simples.csv"
    caminho.write_text(
        "Estudante,RA,Turma,Trimestre,Arte - Frente Única - AV 1 Obj\n"
        "ALICE,1222,1A,T1,7\n",
        encoding="utf-8",
    )
    return caminho


# ---------------------------------------------------------------------------
# Testes de listar_abas_xlsx
# ---------------------------------------------------------------------------

class TestListarAbas:
    def test_aba_unica_retorna_lista_com_1_elemento(self, xlsx_aba_unica):
        abas = listar_abas_xlsx(xlsx_aba_unica)
        assert abas == ["Notas"]

    def test_multi_aba_retorna_todas(self, xlsx_multi_aba):
        abas = listar_abas_xlsx(xlsx_multi_aba)
        assert abas == ["1A_T1", "1A_T2", "2A_T1", "2A_T2"]

    def test_csv_retorna_lista_vazia(self, csv_simples):
        abas = listar_abas_xlsx(csv_simples)
        assert abas == []

    def test_path_como_string(self, xlsx_multi_aba):
        abas = listar_abas_xlsx(str(xlsx_multi_aba))
        assert len(abas) == 4


# ---------------------------------------------------------------------------
# Testes de carregar_entrada — aba única (retrocompatibilidade)
# ---------------------------------------------------------------------------

class TestCarregarEntradaAbaUnica:
    def test_sem_sheet_name_le_primeira_aba(self, xlsx_aba_unica):
        df = carregar_entrada(xlsx_aba_unica)
        assert len(df) == 1
        assert df.iloc[0]["Estudante"] == "ALICE"

    def test_sheet_name_none_explicito_le_corretamente(self, xlsx_aba_unica):
        df = carregar_entrada(xlsx_aba_unica, sheet_name=None)
        assert len(df) == 1

    def test_sheet_name_correto_le_aba(self, xlsx_aba_unica):
        df = carregar_entrada(xlsx_aba_unica, sheet_name="Notas")
        assert len(df) == 1
        assert df.iloc[0]["RA"] == "1222"

    def test_dataframe_ignorado_sheet_name(self):
        """DataFrame passado diretamente: sheet_name é ignorado silenciosamente."""
        df_in = pd.DataFrame([{"Estudante": "X", "RA": "1", "Turma": "1A",
                                "Trimestre": "T1"}])
        df_out = carregar_entrada(df_in, sheet_name="qualquer")
        assert len(df_out) == 1


# ---------------------------------------------------------------------------
# Testes de carregar_entrada — multi-aba
# ---------------------------------------------------------------------------

class TestCarregarEntradaMultiAba:
    def test_sem_sheet_name_lanca_template_invalido(self, xlsx_multi_aba):
        """Multi-aba sem seleção → erro explícito com lista de abas."""
        with pytest.raises(TemplateInvalidoError) as exc_info:
            carregar_entrada(xlsx_multi_aba)
        msg = str(exc_info.value)
        assert "--aba" in msg
        assert "1A_T1" in msg

    def test_sheet_name_valido_le_aba_correta(self, xlsx_multi_aba):
        df = carregar_entrada(xlsx_multi_aba, sheet_name="2A_T1")
        assert len(df) == 1
        assert df.iloc[0]["Estudante"] == "ALUNO_2A_T1"
        assert df.iloc[0]["Turma"] == "2A"
        assert df.iloc[0]["Trimestre"] == "T1"

    def test_sheet_name_outra_aba(self, xlsx_multi_aba):
        df = carregar_entrada(xlsx_multi_aba, sheet_name="1A_T2")
        assert df.iloc[0]["Estudante"] == "ALUNO_1A_T2"

    def test_sheet_name_inexistente_lanca_template_invalido(self, xlsx_multi_aba):
        with pytest.raises(TemplateInvalidoError) as exc_info:
            carregar_entrada(xlsx_multi_aba, sheet_name="3B_T9")
        msg = str(exc_info.value)
        assert "3B_T9" in msg
        assert "1A_T1" in msg  # lista as abas disponíveis

    def test_multi_aba_dados_corretos_todas_abas(self, xlsx_multi_aba):
        """Cada aba deve ser carregável individualmente com dados corretos."""
        for nome_aba in ["1A_T1", "1A_T2", "2A_T1", "2A_T2"]:
            df = carregar_entrada(xlsx_multi_aba, sheet_name=nome_aba)
            assert df.iloc[0]["Estudante"] == f"ALUNO_{nome_aba}"


# ---------------------------------------------------------------------------
# Testes de carregar_entrada — CSV
# ---------------------------------------------------------------------------

class TestCarregarEntradaCSV:
    def test_csv_sem_sheet_name(self, csv_simples):
        df = carregar_entrada(csv_simples)
        assert len(df) == 1
        assert df.iloc[0]["Estudante"] == "ALICE"

    def test_csv_com_sheet_name_ignorado(self, csv_simples):
        """CSV com sheet_name definido não deve lançar erro — apenas ignora."""
        df = carregar_entrada(csv_simples, sheet_name="qualquer_coisa")
        assert len(df) == 1

    def test_csv_arquivo_nao_encontrado(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            carregar_entrada(tmp_path / "nao_existe.csv")


# ---------------------------------------------------------------------------
# Testes de segurança
# ---------------------------------------------------------------------------

class TestSeguranca:
    def test_nao_processa_todas_abas_automaticamente(self, xlsx_multi_aba):
        """
        Sem --aba, nunca processa dados de abas automaticamente.
        O pipeline deve falhar explicitamente, nunca silenciosamente pegar
        a primeira aba de um workbook anual e processar sem o operador saber.
        """
        with pytest.raises(TemplateInvalidoError):
            carregar_entrada(xlsx_multi_aba)

    def test_mensagem_de_erro_menciona_flag_aba(self, xlsx_multi_aba):
        """A mensagem de erro deve guiar o operador para usar --aba."""
        with pytest.raises(TemplateInvalidoError) as exc_info:
            carregar_entrada(xlsx_multi_aba)
        assert "--aba" in str(exc_info.value)

    def test_workbook_anual_real(self, tmp_path):
        """
        Simula o workbook anual completo (12 abas).
        Sem --aba → erro. Com --aba → dados corretos da aba selecionada.
        """
        wb = openpyxl.Workbook()
        abas = [f"{t}_{tr}" for t in ["1A", "1B", "2A", "2B"] for tr in ["T1", "T2", "T3"]]
        for nome in abas:
            ws = wb.create_sheet(nome)
            ws.append(["Estudante", "RA", "Turma", "Trimestre"])
            ws.append([f"Aluno_{nome}", "999", nome[:2], nome[3:]])
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        caminho = tmp_path / "madan_2026_anual.xlsx"
        wb.save(caminho)

        # Sem aba → erro
        with pytest.raises(TemplateInvalidoError):
            carregar_entrada(caminho)

        # Com aba → correto
        df = carregar_entrada(caminho, sheet_name="2B_T3")
        assert df.iloc[0]["Estudante"] == "Aluno_2B_T3"
        assert df.iloc[0]["Trimestre"] == "T3"
