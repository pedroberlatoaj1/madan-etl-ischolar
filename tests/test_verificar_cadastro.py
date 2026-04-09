from __future__ import annotations

from io import StringIO
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from verificar_cadastro import (
    LeituraPlanilhaError,
    RelatorioTurma,
    ResultadoAluno,
    deduplicar_turmas_plano_b,
    imprimir_relatorio,
    ler_planilha_multi_aba,
    parsear_nome_aba_plano_b,
    salvar_csv,
    verificar_aluno,
)


def _criar_planilha(tmp_path: Path, abas: dict[str, list[list[object]]]) -> Path:
    wb = openpyxl.Workbook()
    primeira = True
    for nome_aba, linhas in abas.items():
        if primeira:
            ws = wb.active
            ws.title = nome_aba
            primeira = False
        else:
            ws = wb.create_sheet(nome_aba)
        for linha in linhas:
            ws.append(linha)
    caminho = tmp_path / "cadastro.xlsx"
    wb.save(caminho)
    wb.close()
    return caminho


class ClienteFake:
    def __init__(self, buscar_resultados, listar_resultados):
        self._buscar_resultados = list(buscar_resultados)
        self._listar_resultados = list(listar_resultados)
        self.listar_chamadas = []

    def buscar_aluno(self, *, ra):
        return self._buscar_resultados.pop(0)

    def listar_matriculas(self, **kwargs):
        self.listar_chamadas.append(kwargs)
        return self._listar_resultados.pop(0)


def _resp_busca(sucesso=True, status_code=200, mensagem="ok", dados=None):
    return SimpleNamespace(
        sucesso=sucesso,
        status_code=status_code,
        mensagem=mensagem,
        dados=dados if dados is not None else {"dados": {"id_aluno": 10}},
    )


def _resp_lista(
    sucesso,
    id_matricula=None,
    mensagem="",
    dados=None,
    rastreabilidade=None,
):
    return SimpleNamespace(
        sucesso=sucesso,
        status_code=200,
        mensagem=mensagem,
        dados=dados,
        id_matricula_resolvido=id_matricula,
        rastreabilidade=rastreabilidade or {},
        transitorio=False,
    )


def test_ler_planilha_multi_aba_por_cabecalho(tmp_path: Path):
    caminho = _criar_planilha(
        tmp_path,
        {
            "1A": [
                ["Estudante", "RA", "Turma"],
                ["ALICE", 1222, None],
                ["BRUNO", 1333, None],
            ]
        },
    )

    leitura = ler_planilha_multi_aba(str(caminho))

    assert leitura.avisos == []
    assert leitura.turmas["1A"] == [
        {"ra": "1222", "nome": "ALICE"},
        {"ra": "1333", "nome": "BRUNO"},
    ]


def test_ler_planilha_multi_aba_detecta_inversao_consistente(tmp_path: Path):
    caminho = _criar_planilha(
        tmp_path,
        {
            "1A": [
                ["Estudante", "RA"],
                [1222, "ALICE"],
                [1333, "BRUNO"],
            ]
        },
    )

    leitura = ler_planilha_multi_aba(str(caminho))

    assert len(leitura.avisos) == 1
    assert "invertidos" in leitura.avisos[0]
    assert leitura.turmas["1A"][0] == {"ra": "1222", "nome": "ALICE"}


def test_ler_planilha_multi_aba_falha_em_inconsistencia(tmp_path: Path):
    caminho = _criar_planilha(
        tmp_path,
        {
            "1A": [
                ["Estudante", "RA"],
                ["ALICE", 1222],
                [1333, "BRUNO"],
            ]
        },
    )

    with pytest.raises(LeituraPlanilhaError, match="inconsistentes"):
        ler_planilha_multi_aba(str(caminho))


def test_verificar_aluno_ra_nao_encontrado():
    cliente = ClienteFake(
        buscar_resultados=[_resp_busca(sucesso=False, status_code=404, mensagem="Aluno nao encontrado")],
        listar_resultados=[],
    )

    resultado = verificar_aluno(cliente, "1A", "9999", "ALUNO TESTE")

    assert resultado.status == "RA_NAO_ENCONTRADO"
    assert "404" in resultado.detalhe


def test_verificar_aluno_matricula_resolvida_diretamente():
    cliente = ClienteFake(
        buscar_resultados=[_resp_busca(dados={"dados": {"id_aluno": 42}})],
        listar_resultados=[_resp_lista(sucesso=True, id_matricula=7001)],
    )

    resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")

    assert resultado.status == "OK"
    assert resultado.id_aluno == 42
    assert resultado.id_matricula == 7001


def test_verificar_aluno_resolve_por_situacao_cursando():
    cliente = ClienteFake(
        buscar_resultados=[_resp_busca(dados={"dados": {"id_aluno": 42}})],
        listar_resultados=[
            _resp_lista(sucesso=False, id_matricula=None, mensagem="Nenhuma matricula", rastreabilidade={"id_matriculas_extraiados": []}),
            _resp_lista(sucesso=True, id_matricula=7002),
        ],
    )

    resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")

    assert resultado.status == "HEURISTICA"
    assert resultado.id_matricula == 7002
    assert "situacao=cursando" in resultado.heuristica_usada


def test_verificar_aluno_resolve_por_status_matriculado():
    cliente = ClienteFake(
        buscar_resultados=[_resp_busca(dados={"dados": {"id_aluno": 42}})],
        listar_resultados=[
            _resp_lista(
                sucesso=False,
                id_matricula=None,
                mensagem="Resposta retornou multiplos id_matricula distintos",
                rastreabilidade={"id_matriculas_extraiados": [111, 222]},
                dados={
                    "dados": [
                        {"id_matricula": 111, "status_matricula_diario": "TRANCADO"},
                        {"id_matricula": 222, "status_matricula_diario": "MATRICULADO"},
                    ]
                },
            ),
            _resp_lista(sucesso=False, id_matricula=None, mensagem="ambigua"),
            _resp_lista(sucesso=False, id_matricula=None, mensagem="ambigua"),
        ],
    )

    resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")

    assert resultado.status == "HEURISTICA"
    assert resultado.id_matricula == 222
    assert "MATRICULADO" in resultado.heuristica_usada


def test_verificar_aluno_matricula_ambigua_sem_solucao():
    cliente = ClienteFake(
        buscar_resultados=[_resp_busca(dados={"dados": {"id_aluno": 42}})],
        listar_resultados=[
            _resp_lista(
                sucesso=False,
                id_matricula=None,
                mensagem="Resposta retornou multiplos id_matricula distintos",
                rastreabilidade={"id_matriculas_extraiados": [111, 222]},
                dados={"dados": [{"id_matricula": 111}, {"id_matricula": 222}]},
            ),
            _resp_lista(sucesso=False, id_matricula=None, mensagem="ambigua"),
            _resp_lista(sucesso=False, id_matricula=None, mensagem="ambigua"),
        ],
    )

    resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")

    assert resultado.status == "MATRICULA_AMBIGUA"
    assert "[111, 222]" in resultado.detalhe


def test_imprimir_relatorio_basico():
    relatorio = RelatorioTurma(
        turma="1A",
        total=2,
        ok=1,
        heuristica=1,
        falhas=0,
        alunos=[
            ResultadoAluno(turma="1A", ra="1222", nome="ALICE", status="OK", id_aluno=1, id_matricula=10),
            ResultadoAluno(
                turma="1A",
                ra="1333",
                nome="BRUNO",
                status="HEURISTICA",
                id_aluno=2,
                id_matricula=20,
                heuristica_usada="situacao=cursando",
            ),
        ],
    )

    saida = StringIO()
    imprimir_relatorio([relatorio], arquivo=saida)
    conteudo = saida.getvalue()

    assert "VERIFICACAO DE CADASTRO" in conteudo
    assert "TURMA: 1A" in conteudo
    assert "situacao=cursando" in conteudo


def test_salvar_csv(tmp_path: Path):
    relatorio = RelatorioTurma(
        turma="2A",
        total=1,
        ok=0,
        heuristica=0,
        falhas=1,
        alunos=[
            ResultadoAluno(
                turma="2A",
                ra="9999",
                nome="ALUNO FANTASMA",
                status="RA_NAO_ENCONTRADO",
                detalhe="HTTP 404",
            )
        ],
    )

    caminho_csv = tmp_path / "resultado.csv"
    salvar_csv([relatorio], str(caminho_csv))

    conteudo = caminho_csv.read_text(encoding="utf-8-sig")
    assert "Turma,RA,Nome,Status" in conteudo
    assert "ALUNO FANTASMA" in conteudo


# ---------------------------------------------------------------------------
# Plano B — parsear_nome_aba_plano_b
# ---------------------------------------------------------------------------

class TestParsearNomeAba:
    def test_2a_t1(self):
        assert parsear_nome_aba_plano_b("2A_T1") == ("2A", "T1")

    def test_1b_t3(self):
        assert parsear_nome_aba_plano_b("1B_T3") == ("1B", "T3")

    def test_lowercase(self):
        assert parsear_nome_aba_plano_b("2a_t2") == ("2A", "T2")

    def test_espacos_ignorados(self):
        assert parsear_nome_aba_plano_b("  1A_T1  ") == ("1A", "T1")

    def test_notas_retorna_none(self):
        assert parsear_nome_aba_plano_b("Notas") is None

    def test_1a_sem_trimestre_retorna_none(self):
        assert parsear_nome_aba_plano_b("1A") is None

    def test_t4_invalido_retorna_none(self):
        assert parsear_nome_aba_plano_b("2A_T4") is None

    def test_string_vazia_retorna_none(self):
        assert parsear_nome_aba_plano_b("") is None

    def test_serie_dois_digitos_retorna_none(self):
        assert parsear_nome_aba_plano_b("12A_T1") is None


# ---------------------------------------------------------------------------
# Plano B — deduplicar_turmas_plano_b
# ---------------------------------------------------------------------------

class TestDeduplicarTurmasPlanoB:
    def _turmas_12_abas(self) -> dict:
        """Simula 12 abas de um workbook anual completo (4 turmas × 3 trimestres)."""
        alunos = {
            "1A": [{"ra": "100", "nome": "ALICE"}, {"ra": "101", "nome": "BRUNO"}],
            "1B": [{"ra": "200", "nome": "CARLA"}],
            "2A": [{"ra": "300", "nome": "DIEGO"}, {"ra": "301", "nome": "EVA"}],
            "2B": [{"ra": "400", "nome": "FABIO"}],
        }
        turmas = {}
        for turma, lista in alunos.items():
            for tri in ("T1", "T2", "T3"):
                turmas[f"{turma}_{tri}"] = list(lista)
        return turmas

    def test_12_abas_viram_4_turmas(self):
        turmas, fontes, _ = deduplicar_turmas_plano_b(self._turmas_12_abas())
        assert set(turmas.keys()) == {"1A", "1B", "2A", "2B"}

    def test_cada_turma_tem_alunos_sem_duplicatas(self):
        turmas, _, _ = deduplicar_turmas_plano_b(self._turmas_12_abas())
        assert len(turmas["1A"]) == 2
        assert len(turmas["2A"]) == 2

    def test_fontes_por_turma_contem_3_abas(self):
        _, fontes, _ = deduplicar_turmas_plano_b(self._turmas_12_abas())
        assert fontes["1A"] == ["1A_T1", "1A_T2", "1A_T3"]
        assert fontes["2B"] == ["2B_T1", "2B_T2", "2B_T3"]

    def test_avisos_listam_turmas_agregadas(self):
        _, _, avisos = deduplicar_turmas_plano_b(self._turmas_12_abas())
        assert len(avisos) == 4
        assert any("1A" in av for av in avisos)

    def test_ra_extra_em_t2_incluido_uma_vez(self):
        """RA novo que aparece apenas em T2 deve ser incluído."""
        turmas_input = {
            "2A_T1": [{"ra": "300", "nome": "DIEGO"}],
            "2A_T2": [{"ra": "300", "nome": "DIEGO"}, {"ra": "302", "nome": "GABI"}],
            "2A_T3": [{"ra": "300", "nome": "DIEGO"}, {"ra": "302", "nome": "GABI"}],
        }
        turmas, _, _ = deduplicar_turmas_plano_b(turmas_input)
        ras = [a["ra"] for a in turmas["2A"]]
        assert ras.count("300") == 1
        assert ras.count("302") == 1
        assert len(ras) == 2

    def test_tab_legada_nao_e_alterada(self):
        """Aba 'Notas' (legada) deve passar sem alteração."""
        turmas_input = {
            "Notas": [{"ra": "999", "nome": "LEGADO"}],
            "2A_T1": [{"ra": "300", "nome": "DIEGO"}],
        }
        turmas, fontes, _ = deduplicar_turmas_plano_b(turmas_input)
        assert "Notas" in turmas
        assert turmas["Notas"] == [{"ra": "999", "nome": "LEGADO"}]
        assert "Notas" not in fontes  # aba legada não é dedup

    def test_aba_unica_sem_trimestres_sem_aviso(self):
        """Aba Plano B sozinha (sem duplicatas) não gera aviso de merge."""
        turmas_input = {"2A_T1": [{"ra": "300", "nome": "DIEGO"}]}
        _, _, avisos = deduplicar_turmas_plano_b(turmas_input)
        assert avisos == []

    def test_dicionario_vazio(self):
        turmas, fontes, avisos = deduplicar_turmas_plano_b({})
        assert turmas == {}
        assert fontes == {}
        assert avisos == []

    def test_ordem_ra_preservada_de_t1(self):
        """Alunos de T1 devem aparecer antes dos de T2 (primeira ocorrência vence)."""
        turmas_input = {
            "1A_T1": [{"ra": "100", "nome": "A"}, {"ra": "101", "nome": "B"}],
            "1A_T2": [{"ra": "102", "nome": "C"}, {"ra": "100", "nome": "A"}],
        }
        turmas, _, _ = deduplicar_turmas_plano_b(turmas_input)
        ras = [a["ra"] for a in turmas["1A"]]
        assert ras == ["100", "101", "102"]


# ---------------------------------------------------------------------------
# Plano B — ler_planilha_multi_aba com workbook anual
# ---------------------------------------------------------------------------

def _criar_workbook_anual(tmp_path: Path) -> Path:
    """Cria um workbook anual com 6 abas (2 turmas × 3 trimestres)."""
    wb = openpyxl.Workbook()
    primeira = True
    for turma in ("1A", "2A"):
        for tri in ("T1", "T2", "T3"):
            nome_aba = f"{turma}_{tri}"
            if primeira:
                ws = wb.active
                ws.title = nome_aba
                primeira = False
            else:
                ws = wb.create_sheet(nome_aba)
            ws.append(["Estudante", "RA"])
            ws.append(["ALICE", 1111])
            ws.append(["BRUNO", 1222])
    caminho = tmp_path / "anual.xlsx"
    wb.save(caminho)
    wb.close()
    return caminho


class TestLerPlanilhaAnual:
    def test_le_todas_abas_plano_b(self, tmp_path):
        caminho = _criar_workbook_anual(tmp_path)
        leitura = ler_planilha_multi_aba(str(caminho))
        assert set(leitura.turmas.keys()) == {
            "1A_T1", "1A_T2", "1A_T3",
            "2A_T1", "2A_T2", "2A_T3",
        }

    def test_filtro_aba_retorna_apenas_uma(self, tmp_path):
        caminho = _criar_workbook_anual(tmp_path)
        leitura = ler_planilha_multi_aba(str(caminho), aba_filtro="2A_T1")
        assert list(leitura.turmas.keys()) == ["2A_T1"]
        assert len(leitura.turmas["2A_T1"]) == 2

    def test_filtro_aba_case_insensitive(self, tmp_path):
        caminho = _criar_workbook_anual(tmp_path)
        leitura = ler_planilha_multi_aba(str(caminho), aba_filtro="2a_t1")
        assert "2A_T1" in leitura.turmas

    def test_alunos_corretos_em_cada_aba(self, tmp_path):
        caminho = _criar_workbook_anual(tmp_path)
        leitura = ler_planilha_multi_aba(str(caminho))
        assert leitura.turmas["1A_T1"] == [
            {"ra": "1111", "nome": "ALICE"},
            {"ra": "1222", "nome": "BRUNO"},
        ]


# ---------------------------------------------------------------------------
# Plano B — imprimir_relatorio com fontes
# ---------------------------------------------------------------------------

def test_imprimir_relatorio_mostra_fontes_quando_deduplicated():
    rel = RelatorioTurma(
        turma="2A",
        total=2,
        ok=2,
        heuristica=0,
        falhas=0,
        fontes=["2A_T1", "2A_T2", "2A_T3"],
        alunos=[
            ResultadoAluno(turma="2A", ra="300", nome="DIEGO", status="OK"),
            ResultadoAluno(turma="2A", ra="301", nome="EVA", status="OK"),
        ],
    )
    saida = StringIO()
    imprimir_relatorio([rel], arquivo=saida)
    conteudo = saida.getvalue()
    assert "abas: 2A_T1, 2A_T2, 2A_T3" in conteudo


def test_imprimir_relatorio_sem_fontes_comportamento_original():
    """Relatório sem fontes (workbook legado) não deve mostrar '[abas: ...]'."""
    rel = RelatorioTurma(
        turma="1A",
        total=1,
        ok=1,
        alunos=[ResultadoAluno(turma="1A", ra="100", nome="X", status="OK")],
    )
    saida = StringIO()
    imprimir_relatorio([rel], arquivo=saida)
    assert "[abas:" not in saida.getvalue()
