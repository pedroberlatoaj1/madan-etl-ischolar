"""
Testes para gerador_planilhas.py — geração de planilhas wide por turma.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from gerador_planilhas import (
    Aluno,
    TabConfig,
    _sanitizar_nome_aba,
    agrupar_alunos_por_turma,
    carregar_roster_csv,
    construir_cabecalho_wide,
    descobrir_grupos_wide,
    descobrir_tabs_para_turma,
    gerar_planilha_turma,
    gerar_todas_planilhas,
    COLUNAS_FIXAS_WIDE,
    TIPOS_AVALIACAO_WIDE,
    DISCIPLINA_DISPLAY,
    # retrocompatibilidade — ainda exportados
    COLUNAS_IDENTIDADE,
    COLUNAS_NOTA,
    COLUNAS_CONFERENCIA,
    TODAS_COLUNAS,
)
from wide_format_adapter import parsear_coluna_dinamica, mapear_tipo_avaliacao


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def alunos_1a() -> list[Aluno]:
    return [
        Aluno(nome="Alice Silva", ra="1001", turma="1A"),
        Aluno(nome="Bruno Costa", ra="1002", turma="1A"),
        Aluno(nome="Carla Dias",  ra="1003", turma="1A"),
    ]


@pytest.fixture
def alunos_mistos() -> list[Aluno]:
    return [
        Aluno(nome="Alice Silva",     ra="1001", turma="1A"),
        Aluno(nome="Bruno Costa",     ra="1002", turma="1A"),
        Aluno(nome="Daniel Lima",     ra="2001", turma="2B"),
        Aluno(nome="Eva Martins",     ra="2002", turma="2B"),
        Aluno(nome="Fernanda Gomes",  ra="3001", turma="3A"),  # 3ª série — ignorada
    ]


@pytest.fixture
def roster_csv(tmp_path: Path) -> Path:
    csv_path = tmp_path / "roster.csv"
    csv_path.write_text(
        "Nome,RA,Turma\n"
        "Alice Silva,1001,1A\n"
        "Bruno Costa,1002,1A\n"
        "Daniel Lima,2001,2B\n",
        encoding="utf-8",
    )
    return csv_path


# ---------------------------------------------------------------------------
# TestSanitizarNomeAba — sem alterações
# ---------------------------------------------------------------------------

class TestSanitizarNomeAba:
    def test_remove_acentos(self):
        assert _sanitizar_nome_aba("história_F1_José") == "historia_F1_Jose"

    def test_remove_caracteres_proibidos(self):
        assert _sanitizar_nome_aba("mat/F1:prof") == "mat_F1_prof"

    def test_preserva_underscore(self):
        assert _sanitizar_nome_aba("bio_F2_Silva") == "bio_F2_Silva"


# ---------------------------------------------------------------------------
# TestTabConfig — mantido para retrocompatibilidade
# ---------------------------------------------------------------------------

class TestTabConfig:
    def test_nome_aba_com_frente(self):
        tab = TabConfig("matematica", "F2", "Luan", "Luan Schunck")
        assert tab.nome_aba == "matematica_F2_Luan"

    def test_nome_aba_sem_frente(self):
        tab = TabConfig("arte", "", "Lenice", "Lenice Silva")
        assert tab.nome_aba == "arte_Lenice"

    def test_nome_aba_truncado_31_chars(self):
        tab = TabConfig("interpretacao de texto", "F1", "ProfessorNomeMuitoLongo", "Full Name")
        assert len(tab.nome_aba) <= 31

    def test_nome_aba_sem_acentos(self):
        tab = TabConfig("história", "F1", "José", "José da Silva")
        assert "ó" not in tab.nome_aba
        assert "é" not in tab.nome_aba


# ---------------------------------------------------------------------------
# TestDescobrirTabs — sem alterações
# ---------------------------------------------------------------------------

class TestDescobrirTabs:
    def test_turma_1a_tem_tabs(self):
        tabs = descobrir_tabs_para_turma(1, "A")
        assert len(tabs) > 0

    def test_turma_1a_tem_matematica(self):
        tabs = descobrir_tabs_para_turma(1, "A")
        disciplinas = {t.disciplina for t in tabs}
        assert "matematica" in disciplinas

    def test_tabs_deduplicadas(self):
        tabs = descobrir_tabs_para_turma(1, "A")
        nomes = [t.nome_aba for t in tabs]
        assert len(nomes) == len(set(nomes))

    def test_turma_2b_tem_tabs(self):
        tabs = descobrir_tabs_para_turma(2, "B")
        assert len(tabs) > 0

    def test_tabs_ordenadas(self):
        tabs = descobrir_tabs_para_turma(1, "A")
        disciplinas = [t.disciplina for t in tabs]
        assert disciplinas == sorted(disciplinas)


# ---------------------------------------------------------------------------
# TestCarregarRoster — sem alterações
# ---------------------------------------------------------------------------

class TestCarregarRoster:
    def test_carrega_csv_basico(self, roster_csv: Path):
        alunos = carregar_roster_csv(roster_csv)
        assert len(alunos) == 3
        assert alunos[0].nome == "Alice Silva"
        assert alunos[0].ra == "1001"
        assert alunos[0].turma == "1A"

    def test_csv_com_headers_alternativos(self, tmp_path: Path):
        csv_path = tmp_path / "alt.csv"
        csv_path.write_text(
            "Estudante,Registro_Aluno,Sala\n"
            "Maria,5001,2A\n",
            encoding="utf-8",
        )
        alunos = carregar_roster_csv(csv_path)
        assert len(alunos) == 1
        assert alunos[0].nome == "Maria"
        assert alunos[0].ra == "5001"

    def test_csv_sem_coluna_obrigatoria_levanta_erro(self, tmp_path: Path):
        csv_path = tmp_path / "bad.csv"
        csv_path.write_text("Nome,Turma\nAlice,1A\n", encoding="utf-8")
        with pytest.raises(ValueError, match="ra"):
            carregar_roster_csv(csv_path)

    def test_csv_pula_linhas_sem_nome_ou_ra(self, tmp_path: Path):
        csv_path = tmp_path / "sparse.csv"
        csv_path.write_text(
            "Nome,RA,Turma\nAlice,1001,1A\n,,1A\nBruno,,1A\n",
            encoding="utf-8",
        )
        alunos = carregar_roster_csv(csv_path)
        assert len(alunos) == 1


# ---------------------------------------------------------------------------
# TestAgruparPorTurma — sem alterações
# ---------------------------------------------------------------------------

class TestAgruparPorTurma:
    def test_agrupa_corretamente(self, alunos_mistos: list[Aluno]):
        grupos = agrupar_alunos_por_turma(alunos_mistos)
        assert "1A" in grupos
        assert "2B" in grupos
        assert len(grupos["1A"]) == 2
        assert len(grupos["2B"]) == 2

    def test_alunos_ordenados_por_nome(self, alunos_1a: list[Aluno]):
        desordenados = list(reversed(alunos_1a))
        grupos = agrupar_alunos_por_turma(desordenados)
        nomes = [a.nome for a in grupos["1A"]]
        assert nomes == sorted(nomes)


# ---------------------------------------------------------------------------
# TestDescobrirGruposWide — NOVO
# ---------------------------------------------------------------------------

class TestDescobrirGruposWide:
    def test_turma_1a_retorna_grupos(self):
        grupos = descobrir_grupos_wide(1, "A")
        assert len(grupos) > 0

    def test_cada_grupo_e_tupla_dois_elementos(self):
        grupos = descobrir_grupos_wide(1, "A")
        for g in grupos:
            assert isinstance(g, tuple)
            assert len(g) == 2

    def test_frente_display_comeca_com_Frente(self):
        grupos = descobrir_grupos_wide(1, "A")
        for _, frente_display in grupos:
            assert frente_display.startswith("Frente "), (
                f"Frente display inválida: {frente_display!r}"
            )

    def test_disciplina_com_multiplas_frentes(self):
        """Física na turma 1A tem múltiplos professores → deve gerar múltiplas frentes."""
        grupos = descobrir_grupos_wide(1, "A")
        frentes_fis = [f for d, f in grupos if d == "Física"]
        assert len(frentes_fis) >= 2, (
            f"Física deve ter pelo menos 2 frentes na turma 1A, "
            f"encontradas: {frentes_fis}"
        )

    def test_disciplina_unica_usa_frente_unica(self):
        """Inglês tem 1 professor (todas as turmas) → Frente Única."""
        grupos = descobrir_grupos_wide(1, "A")
        frentes_ingles = [f for d, f in grupos if d == "Inglês"]
        assert frentes_ingles == ["Frente Única"]

    def test_grupos_ordenados_por_disciplina(self):
        grupos = descobrir_grupos_wide(1, "A")
        disciplinas = [d for d, _ in grupos]
        assert disciplinas == sorted(disciplinas)

    def test_disciplinas_usam_display_names(self):
        """Disciplinas devem ser nomes de exibição, não slugs."""
        grupos = descobrir_grupos_wide(1, "A")
        disciplinas = {d for d, _ in grupos}
        valores_display = set(DISCIPLINA_DISPLAY.values())
        for disc in disciplinas:
            assert disc in valores_display, (
                f"Disciplina {disc!r} não está em DISCIPLINA_DISPLAY"
            )

    def test_turma_2b_retorna_grupos(self):
        grupos = descobrir_grupos_wide(2, "B")
        assert len(grupos) > 0

    def test_filosofia_aparece_nos_grupos_do_primeiro_ano(self):
        grupos = descobrir_grupos_wide(1, "A")
        disciplinas = {d for d, _ in grupos}
        assert "Filosofia" in disciplinas


# ---------------------------------------------------------------------------
# TestConstruirCabecalhoWide — NOVO
# ---------------------------------------------------------------------------

class TestConstruirCabecalhoWide:
    @pytest.fixture
    def grupos_simples(self) -> list[tuple[str, str]]:
        return [
            ("Matemática", "Frente A"),
            ("Matemática", "Frente B"),
            ("Física",     "Frente Única"),
        ]

    def test_comeca_com_colunas_fixas(self, grupos_simples):
        cab = construir_cabecalho_wide(grupos_simples)
        assert cab[:4] == COLUNAS_FIXAS_WIDE

    def test_largura_total_correta(self, grupos_simples):
        cab = construir_cabecalho_wide(grupos_simples)
        n_tipos_sem_rec_final = len(TIPOS_AVALIACAO_WIDE) - 1
        n_esperado = len(COLUNAS_FIXAS_WIDE) + len(grupos_simples) * n_tipos_sem_rec_final
        assert len(cab) == n_esperado

    def test_largura_total_correta_no_t3_com_recuperacao_final(self, grupos_simples):
        cab = construir_cabecalho_wide(grupos_simples, incluir_recuperacao_final=True)
        n_esperado = len(COLUNAS_FIXAS_WIDE) + len(grupos_simples) * len(TIPOS_AVALIACAO_WIDE)
        assert len(cab) == n_esperado

    def test_colunas_dinamicas_match_regex(self, grupos_simples):
        """Todas as colunas dinâmicas devem ser reconhecidas por parsear_coluna_dinamica."""
        cab = construir_cabecalho_wide(grupos_simples)
        fixas = set(COLUNAS_FIXAS_WIDE)
        for col in cab:
            if col in fixas:
                continue
            resultado = parsear_coluna_dinamica(col)
            assert resultado is not None, (
                f"parsear_coluna_dinamica não reconheceu: {col!r}"
            )

    def test_tipo_avaliacao_mapeavel(self, grupos_simples):
        """O tipo extraído de cada coluna deve ser mapeável para coluna canônica."""
        cab = construir_cabecalho_wide(grupos_simples)
        fixas = set(COLUNAS_FIXAS_WIDE)
        for col in cab:
            if col in fixas:
                continue
            parsed = parsear_coluna_dinamica(col)
            assert parsed is not None
            mapeado = mapear_tipo_avaliacao(parsed.tipo_avaliacao)
            assert mapeado is not None, (
                f"mapear_tipo_avaliacao falhou para tipo {parsed.tipo_avaliacao!r} "
                f"extraído de {col!r}"
            )

    def test_round_trip_disciplina_e_frente(self, grupos_simples):
        """Disciplina e frente extraídas devem bater com as fornecidas."""
        cab = construir_cabecalho_wide(grupos_simples)
        fixas = set(COLUNAS_FIXAS_WIDE)

        extraidos: set[tuple[str, str]] = set()
        for col in cab:
            if col in fixas:
                continue
            p = parsear_coluna_dinamica(col)
            assert p is not None
            extraidos.add((p.disciplina, p.frente))

        grupos_set = {(d, f) for d, f in grupos_simples}
        assert extraidos == grupos_set

    def test_turma_real_1a_colunas_validas(self):
        """Integração: header gerado para turma 1A deve passar 100% no adapter."""
        grupos = descobrir_grupos_wide(1, "A")
        cab = construir_cabecalho_wide(grupos)
        fixas = set(COLUNAS_FIXAS_WIDE)
        invalidas = []
        for col in cab:
            if col in fixas:
                continue
            p = parsear_coluna_dinamica(col)
            if p is None:
                invalidas.append(col)
            elif mapear_tipo_avaliacao(p.tipo_avaliacao) is None:
                invalidas.append(col)
        assert invalidas == [], f"Colunas inválidas: {invalidas}"

    def test_sem_grupos_retorna_so_fixas(self):
        cab = construir_cabecalho_wide([])
        assert cab == COLUNAS_FIXAS_WIDE

    def test_t3_inclui_recuperacao_final_quando_solicitado(self, grupos_simples):
        cab = construir_cabecalho_wide(grupos_simples, incluir_recuperacao_final=True)
        assert "Matemática - Frente A - Recuperação Final" in cab

    def test_t1_t2_nao_incluem_recuperacao_final_por_padrao(self, grupos_simples):
        cab = construir_cabecalho_wide(grupos_simples)
        assert "Matemática - Frente A - Recuperação Final" not in cab


# ---------------------------------------------------------------------------
# TestGerarPlanilha — atualizado para formato wide
# ---------------------------------------------------------------------------

class TestGerarPlanilha:
    def test_gera_arquivo_xlsx(self, alunos_1a: list[Aluno], tmp_path: Path):
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        assert filepath.exists()
        assert filepath.suffix == ".xlsx"
        assert "1A" in filepath.name

    def test_contem_aba_notas(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Formato wide: exatamente 1 aba chamada 'Notas'."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        assert wb.sheetnames == ["Notas"]
        wb.close()

    def test_colunas_fixas_no_inicio(self, alunos_1a: list[Aluno], tmp_path: Path):
        """As 4 primeiras colunas devem ser as fixas."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        primeiras = [ws.cell(row=1, column=i).value for i in range(1, 5)]
        assert primeiras == COLUNAS_FIXAS_WIDE
        wb.close()

    def test_colunas_dinamicas_match_adapter(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Todas as colunas dinâmicas devem ser reconhecidas pelo adapter."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        n_cols = ws.max_column
        headers = [ws.cell(row=1, column=i).value for i in range(1, n_cols + 1)]
        fixas = set(COLUNAS_FIXAS_WIDE)
        invalidas = [
            h for h in headers
            if h not in fixas and parsear_coluna_dinamica(h) is None
        ]
        assert invalidas == []
        wb.close()

    def test_colunas_com_mais_que_fixas(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Deve haver mais colunas que apenas as 4 fixas."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        assert ws.max_column > len(COLUNAS_FIXAS_WIDE)
        wb.close()

    def test_dados_pre_preenchidos(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Colunas fixas devem estar pré-preenchidas na linha 2."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        assert ws.cell(row=2, column=1).value == "Alice Silva"  # Estudante
        assert ws.cell(row=2, column=2).value == "1001"         # RA
        assert ws.cell(row=2, column=3).value == "1A"           # Turma
        assert ws.cell(row=2, column=4).value == "T1"           # Trimestre
        wb.close()

    def test_trimestre_preenchido_em_todos_alunos(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Coluna Trimestre deve estar preenchida para todos os alunos."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        for r in range(2, len(alunos_1a) + 2):
            assert ws.cell(row=r, column=4).value == "T1"
        wb.close()

    def test_numero_de_linhas_correto(self, alunos_1a: list[Aluno], tmp_path: Path):
        """Deve ter 1 linha de cabeçalho + 1 linha por aluno."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        assert ws.max_row == 1 + len(alunos_1a)
        wb.close()

    def test_colunas_nota_ficam_em_branco(self, alunos_1a: list[Aluno], tmp_path: Path):
        """As colunas dinâmicas de nota devem estar vazias (preenchimento pelos professores)."""
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        # Coluna 5 em diante = notas
        for r in range(2, len(alunos_1a) + 2):
            for c in range(5, min(ws.max_column + 1, 10)):  # verifica primeiras 5 notas
                assert ws.cell(row=r, column=c).value is None
        wb.close()

    def test_serie_3_levanta_erro(self, tmp_path: Path):
        alunos = [Aluno(nome="Test", ra="9999", turma="3A")]
        with pytest.raises(ValueError, match="não é suportada"):
            gerar_planilha_turma("3A", "T1", 2026, alunos, tmp_path)

    def test_turma_invalida_levanta_erro(self, tmp_path: Path):
        alunos = [Aluno(nome="Test", ra="9999", turma="XZ")]
        with pytest.raises(ValueError):
            gerar_planilha_turma("XZ", "T1", 2026, alunos, tmp_path)

    def test_planilha_1a_round_trip_adapter(self, alunos_1a: list[Aluno], tmp_path: Path):
        """
        Integração end-to-end: planilha gerada pelo gerador deve ser
        detectada como wide_novo e despivotada sem erros.
        """
        import pandas as pd
        from wide_format_adapter import (
            detectar_formato,
            despivotar_dataframe,
            FORMATO_WIDE_NOVO,
        )

        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        df = pd.read_excel(str(filepath))

        assert detectar_formato(list(df.columns)) == FORMATO_WIDE_NOVO

        df_virtual = despivotar_dataframe(df)
        grupos = descobrir_grupos_wide(1, "A")
        # 3 alunos × N grupos de disciplina/frente
        assert len(df_virtual) == len(alunos_1a) * len(grupos)

    def test_planilha_t3_inclui_coluna_recuperacao_final(self, alunos_1a: list[Aluno], tmp_path: Path):
        filepath = gerar_planilha_turma("1A", "T3", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert any(str(h).endswith("Recuperação Final") for h in headers)
        wb.close()

    def test_planilha_t1_nao_inclui_coluna_recuperacao_final(self, alunos_1a: list[Aluno], tmp_path: Path):
        filepath = gerar_planilha_turma("1A", "T1", 2026, alunos_1a, tmp_path)
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb["Notas"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert not any(str(h).endswith("Recuperação Final") for h in headers)
        wb.close()


# ---------------------------------------------------------------------------
# TestGerarTodas — sem alterações na lógica
# ---------------------------------------------------------------------------

class TestGerarTodas:
    def test_gera_apenas_series_suportadas(self, alunos_mistos: list[Aluno], tmp_path: Path):
        arquivos = gerar_todas_planilhas("T1", 2026, alunos_mistos, tmp_path)
        nomes = {f.name for f in arquivos}
        assert "1A_T1_2026.xlsx" in nomes
        assert "2B_T1_2026.xlsx" in nomes
        assert not any("3A" in n for n in nomes)

    def test_gera_numero_correto_de_arquivos(self, alunos_mistos: list[Aluno], tmp_path: Path):
        arquivos = gerar_todas_planilhas("T1", 2026, alunos_mistos, tmp_path)
        assert len(arquivos) == 2  # 1A e 2B (3A ignorada)
