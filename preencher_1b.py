from openpyxl import load_workbook

students_1b = [
    (1316, "ALICE MENEGHELLI DUARTE"),
    (1261, "ANA CLARA RANGEL FARES"),
    (1436, "ANTONIO VON MUHLEN"),
    (1416, "ARTHUR COVRE DAMIANI"),
    (1318, "ARTHUR PARPAIOLA DE MENEZES"),
    (1243, "BEATRIZ BASTOS RODRIGUES DE FREITAS"),
    (1434, "BERNARDO GAMA FERREIRA"),
    (1519, "CAIO ESTEVES MARTINS"),
    (1315, "DAVI DEL PIERO NUNES"),
    (1309, "DAVI GABRIEL TOSE"),
    (1409, "DIEGO SPELTA DE OLIVEIRA"),
    (1306, "GABRIEL MENDES GAMA"),
    (1232, "GABRIEL ZON LYRIO"),
    (1310, "HENRIQUE BONESI LELLIS"),
    (1321, "ISABELA ROSSI REIS"),
    (1511, "ISABELLA MARBA MIRANDA"),
    (1262, "JOÃO HENRIQUE ALVES COTRIM"),
    (1392, "JULIA DE SOUZA NETTO DAROZ"),
    (1257, "LARA DE ALMEIDA JUNQUEIRA"),
    (1197, "LAURA DOS SANTOS SILVEROL GAURINK DIAS"),
    (1229, "LIGIA SAYURI YOSHIKAWA DE FARIA"),
    (1414, "LUCCA DEPIANTI PINHEIRO"),
    (1213, "LUISA DA ROS LOURENCO MARTINS"),
    (1223, "MANUELA OTTONI CASTRO"),
    (1357, "MARIA EDUARDA FIRME DE OLIVEIRA"),
    (1394, "MARIA FERNANDA CARVALHO DE OLIVEIRA"),
    (1425, "MARIA FERNANDA MAGGIONI AGOSTINHO"),
    (1245, "MARIA FERNANDA OLIVEIRA COMPANHONI"),
    (1212, "MARIANA MOSCON MARÇAL"),
    (1211, "MELINA DA ROS COUTINHO"),
    (1255, "MURILO HILAL SCHMIDT"),
    (1374, "NATALIE LAVIGNE ABILIO DELPUDO"),
    (1248, "NOEMI FREITAS CUNHA"),
    (1237, "OLÍVIA MARTINS PEIXOTO"),
    (1238, "PEDRO LIMA BREDA"),
    (1294, "PEDRO NATAL DE ALMEIDA"),
    (1287, "RAFAEL DE OLIVEIRA SIMÕES"),
    (1290, "RAFAELA DO CARMO STARLING"),
    (1296, "SOFIA BAZILIO FREITAS"),
    (1210, "VINÍCIUS PALASSI TALLON NETTO"),
    (1349, "VITOR BRAGA MIRANDA"),
]

filename = "planilha_madan_2026_preenchida_alunos.xlsx"
wb = load_workbook(filename)
ws = wb["1B"]

for i, (ra, nome) in enumerate(students_1b, start=2):
    ws.cell(row=i, column=1, value=ra)
    ws.cell(row=i, column=2, value=nome)

wb.save(filename)
print(f"Saved. Wrote {len(students_1b)} students to sheet '1B'.")

# Verify
wb2 = load_workbook(filename)
ws2 = wb2["1B"]
data_rows = sum(1 for row in ws2.iter_rows(min_row=2) if any(c.value for c in row))
print(f"Verification: {data_rows} data rows found in sheet '1B'.")
