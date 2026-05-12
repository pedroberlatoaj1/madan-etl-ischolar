"""
Testes unitários para verificar_cadastro.py.

Testa a lógica de verificação de cadastro sem fazer chamadas reais à API.
"""

from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass, field
from typing import Any, Optional
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from verificar_cadastro import (
    ResultadoAluno,
    RelatorioTurma,
    ler_planilha_multi_aba,
    verificar_aluno,
    imprimir_relatorio,
    salvar_csv,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def planilha_temp(tmp_path):
    """Cria planilha temporária com 2 abas e alunos de teste."""
    caminho = tmp_path / "teste_cadastro.xlsx"
    wb = openpyxl.Workbook()

    # Aba 1A — formato real: col A = Estudante (nome), col B = RA
    ws1 = wb.active
    ws1.title = "1A"
    ws1["A1"] = "Estudante"
    ws1["B1"] = "RA"
    ws1["A2"] = "ALICE BARCELOS LINS"
    ws1["B2"] = 1222
    ws1["A3"] = "ALICE DE MEDEIROS GARCIA"
    ws1["B3"] = 1239

    # Aba 2B
    ws2 = wb.create_sheet("2B")
    ws2["A1"] = "Estudante"
    ws2["B1"] = "RA"
    ws2["A2"] = "ALICE PATERLINI AGRELLO"
    ws2["B2"] = 940

    # Aba vazia (sem alunos)
    ws3 = wb.create_sheet("Vazia")
    ws3["A1"] = "Estudante"
    ws3["B1"] = "RA"

    wb.save(caminho)
    return str(caminho)


def _mock_cliente_ok():
    """Cliente mock que retorna sucesso para tudo."""
    cliente = MagicMock()

    # buscar_aluno retorna sucesso com id_aluno
    busca_result = MagicMock()
    busca_result.sucesso = True
    busca_result.status_code = 200
    busca_result.dados = {"id_aluno": 42}
    cliente.buscar_aluno.return_value = busca_result

    # listar_matriculas retorna sucesso com id_matricula
    lista_result = MagicMock()
    lista_result.sucesso = True
    lista_result.id_matricula_resolvido = 1001
    lista_result.rastreabilidade = {}
    cliente.listar_matriculas.return_value = lista_result

    return cliente


def _mock_cliente_ra_falha():
    """Cliente mock que falha em buscar_aluno."""
    cliente = MagicMock()

    busca_result = MagicMock()
    busca_result.sucesso = False
    busca_result.status_code = 404
    busca_result.mensagem = "Aluno não encontrado"
    busca_result.transitorio = False
    cliente.buscar_aluno.return_value = busca_result

    return cliente


def _mock_cliente_matricula_inacessivel():
    """Cliente mock que encontra id_aluno mas matrícula falha."""
    cliente = MagicMock()

    # buscar_aluno OK
    busca_result = MagicMock()
    busca_result.sucesso = True
    busca_result.status_code = 200
    busca_result.dados = {"id_aluno": 42}
    cliente.buscar_aluno.return_value = busca_result

    # listar_matriculas falha (0 resultados, sem heurística CURSANDO)
    lista_result = MagicMock()
    lista_result.sucesso = False
    lista_result.id_matricula_resolvido = None
    lista_result.rastreabilidade = {"id_matriculas_extraiados": []}
    lista_result.mensagem = "Nenhuma matrícula encontrada"
    cliente.listar_matriculas.return_value = lista_result

    return cliente


def _mock_cliente_heuristica():
    """Cliente mock que falha sem filtro mas resolve com CURSANDO."""
    cliente = MagicMock()

    # buscar_aluno OK
    busca_result = MagicMock()
    busca_result.sucesso = True
    busca_result.status_code = 200
    busca_result.dados = {"id_aluno": 42}
    cliente.buscar_aluno.return_value = busca_result

    # listar_matriculas: falha sem filtro, sucesso com situacao
    def listar_side_effect(*, id_aluno, resolver_id_matricula=True, situacao=None, **kw):
        if situacao is None:
            result = MagicMock()
            result.sucesso = False
            result.id_matricula_resolvido = None
            result.rastreabilidade = {"id_matriculas_extraiados": []}
            result.mensagem = "Nenhuma matrícula"
            return result
        else:
            result = MagicMock()
            result.sucesso = True
            result.id_matricula_resolvido = 2002
            result.rastreabilidade = {}
            return result

    cliente.listar_matriculas.side_effect = listar_side_effect

    return cliente


# ---------------------------------------------------------------------------
# Testes de leitura de planilha
# ---------------------------------------------------------------------------

class TestLerPlanilhaMultiAba:
    def test_le_todas_abas(self, planilha_temp):
        leitura = ler_planilha_multi_aba(planilha_temp)
        turmas = leitura.turmas
        assert "1A" in turmas
        assert "2B" in turmas
        assert len(turmas["1A"]) == 2
        assert len(turmas["2B"]) == 1
        # Aba vazia não aparece
        assert "Vazia" not in turmas

    def test_filtro_por_aba(self, planilha_temp):
        leitura = ler_planilha_multi_aba(planilha_temp, aba_filtro="2B")
        turmas = leitura.turmas
        assert "2B" in turmas
        assert "1A" not in turmas

    def test_filtro_case_insensitive(self, planilha_temp):
        leitura = ler_planilha_multi_aba(planilha_temp, aba_filtro="2b")
        turmas = leitura.turmas
        assert "2B" in turmas

    def test_dados_aluno(self, planilha_temp):
        leitura = ler_planilha_multi_aba(planilha_temp)
        aluno = leitura.turmas["1A"][0]
        assert aluno["ra"] == "1222"
        assert aluno["nome"] == "ALICE BARCELOS LINS"


# ---------------------------------------------------------------------------
# Testes de verificação de aluno
# ---------------------------------------------------------------------------

class TestVerificarAluno:
    def test_aluno_ok(self):
        cliente = _mock_cliente_ok()
        resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")
        assert resultado.status == "OK"
        assert resultado.id_aluno == 42
        assert resultado.id_matricula == 1001

    def test_ra_nao_encontrado(self):
        cliente = _mock_cliente_ra_falha()
        resultado = verificar_aluno(cliente, "1A", "9999", "FANTASMA")
        assert resultado.status == "RA_NAO_ENCONTRADO"

    def test_matricula_inacessivel(self):
        cliente = _mock_cliente_matricula_inacessivel()
        resultado = verificar_aluno(cliente, "1A", "1234", "JOSE")
        assert resultado.status == "MATRICULA_INACESSIVEL"
        assert resultado.id_aluno == 42

    def test_heuristica_cursando(self):
        cliente = _mock_cliente_heuristica()
        resultado = verificar_aluno(cliente, "2A", "1065", "AGEU")
        assert resultado.status == "HEURISTICA"
        assert resultado.id_matricula == 2002
        assert "cursando" in resultado.heuristica_usada.lower()

    def test_erro_rede(self):
        cliente = MagicMock()
        cliente.buscar_aluno.side_effect = ConnectionError("timeout")
        resultado = verificar_aluno(cliente, "1A", "1222", "ALICE")
        assert resultado.status == "ERRO_REDE"


# ---------------------------------------------------------------------------
# Testes de relatório
# ---------------------------------------------------------------------------

class TestRelatorio:
    def test_imprimir_relatorio(self, capsys):
        rel = RelatorioTurma(
            turma="1A", total=3, ok=2, heuristica=0, falhas=1,
            alunos=[
                ResultadoAluno(turma="1A", ra="1222", nome="ALICE", status="OK",
                               id_aluno=42, id_matricula=1001),
                ResultadoAluno(turma="1A", ra="1239", nome="MARIA", status="OK",
                               id_aluno=43, id_matricula=1002),
                ResultadoAluno(turma="1A", ra="9999", nome="FANTASMA",
                               status="RA_NAO_ENCONTRADO",
                               detalhe="HTTP 404"),
            ],
        )
        imprimir_relatorio([rel])
        captured = capsys.readouterr()
        assert "TURMA: 1A" in captured.out
        assert "FANTASMA" in captured.out
        assert "RA_NAO_ENCONTRADO" in captured.out
        assert "OK: 2" in captured.out

    def test_salvar_csv(self, tmp_path):
        rel = RelatorioTurma(
            turma="2A", total=1, ok=1, heuristica=0, falhas=0,
            alunos=[
                ResultadoAluno(turma="2A", ra="1065", nome="AGEU", status="OK",
                               id_aluno=10, id_matricula=100),
            ],
        )
        csv_path = str(tmp_path / "resultado.csv")
        salvar_csv([rel], csv_path)

        import csv
        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) == 2  # header + 1 aluno
        assert rows[0][0] == "Turma"
        assert rows[1][0] == "2A"
        assert rows[1][1] == "1065"
